
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_file
from werkzeug.security import generate_password_hash, check_password_hash
import sqlite3
import os
import secrets
from datetime import datetime, timezone
from zoneinfo import ZoneInfo
import pandas as pd
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import shutil
import json
import zipfile
from pathlib import Path
import arabic_reshaper
from bidi.algorithm import get_display
import uuid
import requests

app = Flask(__name__)
app.secret_key = os.environ.get("SESSION_SECRET")

# Generate CSRF token for forms
def generate_csrf_token():
    if 'csrf_token' not in session:
        session['csrf_token'] = secrets.token_hex(16)
    return session['csrf_token']

def validate_csrf_token(token):
    return token and session.get('csrf_token') == token

app.jinja_env.globals['csrf_token'] = generate_csrf_token

# CSRF Protection decorator
def csrf_protect(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if request.method == 'POST':
            token = request.form.get('csrf_token')
            if not validate_csrf_token(token):
                flash('رمز الأمان غير صحيح. يرجى المحاولة مرة أخرى.', 'error')
                return redirect(request.referrer or url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

# Timezone configuration for Palestine (GMT+3)
app.config['LOCAL_TIMEZONE'] = os.getenv('APP_TIMEZONE', 'Asia/Gaza')

# حد أقصى لحجم الملفات المرفوعة (حماية من DoS)
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100 MB

# Jinja filter to convert UTC timestamps to local time (12-hour format)
def local_dt(timestamp_str, format_str='%Y-%m-%d %I:%M %p'):
    """Convert UTC timestamp string to local timezone and format it in 12-hour format"""
    if not timestamp_str:
        return '-'
    
    try:
        # Parse the timestamp (from SQLite CURRENT_TIMESTAMP)
        if isinstance(timestamp_str, str):
            # Remove microseconds if present
            if '.' in timestamp_str:
                timestamp_str = timestamp_str.split('.')[0]
            dt = datetime.fromisoformat(timestamp_str.replace(' ', 'T'))
        else:
            dt = timestamp_str
        
        # If datetime is naive, assume it's UTC
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        
        # Convert to local timezone
        local_tz = ZoneInfo(app.config['LOCAL_TIMEZONE'])
        local_dt = dt.astimezone(local_tz)
        
        # Format time and translate AM/PM to Arabic
        formatted_time = local_dt.strftime(format_str)
        formatted_time = formatted_time.replace('AM', 'ص').replace('PM', 'م')
        
        return formatted_time
    except Exception:
        return str(timestamp_str)

# Register the filter
app.jinja_env.filters['local_dt'] = local_dt

# Secure database connection with foreign key constraints enabled
def get_db_connection():
    """Get database connection with foreign key constraints enabled"""
    conn = sqlite3.connect('database.db')
    conn.execute("PRAGMA foreign_keys=ON")  # Enable foreign key constraints
    return conn

# Permission checking decorator
from functools import wraps

def require_permission(permission_name):
    """Decorator to require specific permission for route access"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                flash('يجب تسجيل الدخول أولاً', 'error')
                return redirect(url_for('login'))
            
            if not has_permission(session['user_id'], permission_name):
                flash(f'ليس لديك صلاحية للوصول لهذه الصفحة', 'error')
                return redirect(url_for('dashboard'))
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def require_login(f):
    """Decorator to require login for route access"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('يجب تسجيل الدخول أولاً', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Database initialization
def init_db():
    conn = get_db_connection()
    c = conn.cursor()
    
    # Users table
    c.execute('''CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL,
        is_admin INTEGER DEFAULT 0,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        session_token TEXT DEFAULT NULL
    )''')
    
    # إضافة عمود session_token للجداول الموجودة (للتوافق مع النسخة القديمة)
    try:
        c.execute('ALTER TABLE users ADD COLUMN session_token TEXT DEFAULT NULL')
    except sqlite3.OperationalError:
        # العمود موجود بالفعل
        pass
    
    # Citizens data table
    c.execute('''CREATE TABLE IF NOT EXISTS citizens (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        full_name TEXT NOT NULL,
        national_id TEXT UNIQUE NOT NULL,
        phone TEXT NOT NULL,
        status TEXT NOT NULL,
        family_members INTEGER NOT NULL,
        address TEXT NOT NULL,
        notes TEXT,
        added_by TEXT NOT NULL,
        assigned_to TEXT DEFAULT NULL,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')
    
    # إضافة عمود assigned_to للجداول الموجودة (للتوافق مع النسخة القديمة)
    try:
        c.execute('ALTER TABLE citizens ADD COLUMN assigned_to TEXT DEFAULT NULL')
    except sqlite3.OperationalError:
        # العمود موجود بالفعل
        pass
    
    # Settings table
    c.execute('''CREATE TABLE IF NOT EXISTS settings (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        site_name TEXT DEFAULT 'نظام البيانات الحديث',
        site_status TEXT DEFAULT 'active',
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')
    
    # Dynamic fields table
    c.execute('''CREATE TABLE IF NOT EXISTS dynamic_fields (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        field_name TEXT UNIQUE NOT NULL,
        field_label TEXT NOT NULL,
        field_type TEXT NOT NULL,
        field_options TEXT,
        is_required INTEGER DEFAULT 0,
        validation_rules TEXT,
        field_order INTEGER DEFAULT 0,
        is_active INTEGER DEFAULT 1,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')
    
    # Dynamic field values table
    c.execute('''CREATE TABLE IF NOT EXISTS dynamic_field_values (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        citizen_id INTEGER NOT NULL,
        field_id INTEGER NOT NULL,
        field_value TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (citizen_id) REFERENCES citizens(id) ON DELETE CASCADE,
        FOREIGN KEY (field_id) REFERENCES dynamic_fields(id) ON DELETE CASCADE
    )''')
    
    # Permissions table للصلاحيات المتقدمة
    c.execute('''CREATE TABLE IF NOT EXISTS permissions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE NOT NULL,
        description TEXT NOT NULL,
        category TEXT NOT NULL,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')
    
    # User permissions table لربط المستخدمين بالصلاحيات
    c.execute('''CREATE TABLE IF NOT EXISTS user_permissions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        permission_id INTEGER NOT NULL,
        granted_by INTEGER NOT NULL,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE,
        FOREIGN KEY (permission_id) REFERENCES permissions(id) ON DELETE CASCADE,
        FOREIGN KEY (granted_by) REFERENCES users(id) ON DELETE CASCADE,
        UNIQUE(user_id, permission_id)
    )''')
    
    # Materials table لتخزين أنواع المواد المختلفة
    c.execute('''CREATE TABLE IF NOT EXISTS materials (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE NOT NULL,
        description TEXT,
        unit TEXT DEFAULT 'قطعة',
        is_active INTEGER DEFAULT 1,
        created_by INTEGER NOT NULL,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (created_by) REFERENCES users(id) ON DELETE CASCADE
    )''')
    
    # Material distributions table لتسجيل توزيع المواد للمواطنين
    c.execute('''CREATE TABLE IF NOT EXISTS material_distributions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        citizen_id INTEGER NOT NULL,
        material_id INTEGER NOT NULL,
        quantity INTEGER DEFAULT 1,
        distributed_by INTEGER NOT NULL,
        distribution_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        notes TEXT,
        FOREIGN KEY (citizen_id) REFERENCES citizens(id) ON DELETE CASCADE,
        FOREIGN KEY (material_id) REFERENCES materials(id) ON DELETE CASCADE,
        FOREIGN KEY (distributed_by) REFERENCES users(id) ON DELETE CASCADE
    )''')
    
    # Telegram backup settings table لإعدادات النسخ الاحتياطي التلقائي عبر تيليجرام
    c.execute('''CREATE TABLE IF NOT EXISTS telegram_backup_settings (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        bot_token TEXT DEFAULT NULL,
        chat_id TEXT DEFAULT NULL,
        is_enabled INTEGER DEFAULT 0,
        backup_on_citizen_changes INTEGER DEFAULT 1,
        backup_on_user_changes INTEGER DEFAULT 1,
        backup_on_material_changes INTEGER DEFAULT 1,
        backup_on_settings_changes INTEGER DEFAULT 1,
        backup_on_permission_changes INTEGER DEFAULT 1,
        last_backup_sent TIMESTAMP DEFAULT NULL,
        backup_file_name_pattern TEXT DEFAULT 'backup_{timestamp}',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')
    
    # Create default admin user
    c.execute("SELECT * FROM users WHERE username = 'admin'")
    if not c.fetchone():
        admin_password = generate_password_hash('admin123')
        c.execute("INSERT INTO users (username, password, is_admin) VALUES (?, ?, 1)", 
                 ('admin', admin_password))
    
    # Create default settings
    c.execute("SELECT * FROM settings")
    if not c.fetchone():
        c.execute("INSERT INTO settings (site_name, site_status) VALUES (?, ?)", 
                 ('نظام البيانات الحديث', 'active'))
    
    # Create default Telegram backup settings
    c.execute("SELECT * FROM telegram_backup_settings")
    if not c.fetchone():
        c.execute("""INSERT INTO telegram_backup_settings 
                     (is_enabled, backup_on_citizen_changes, backup_on_user_changes, 
                      backup_on_material_changes, backup_on_settings_changes, 
                      backup_on_permission_changes, backup_file_name_pattern) 
                     VALUES (0, 1, 1, 1, 1, 1, 'backup_{timestamp}')""")
    
    # Create default permissions
    default_permissions = [
        ('view_citizens', 'عرض بيانات المواطنين', 'البيانات'),
        ('add_citizens', 'إضافة بيانات المواطنين', 'البيانات'),
        ('edit_citizens', 'تعديل بيانات المواطنين', 'البيانات'),
        ('delete_citizens', 'حذف بيانات المواطنين', 'البيانات'),
        ('export_data', 'تصدير البيانات', 'التقارير'),
        ('export_advanced', 'التصدير المتقدم', 'التقارير'),
        ('backup_database', 'إنشاء نسخ احتياطية', 'النسخ الاحتياطية'),
        ('restore_database', 'استعادة النسخ الاحتياطية', 'النسخ الاحتياطية'),
        ('manage_users', 'إدارة المستخدمين', 'الإدارة'),
        ('manage_settings', 'إدارة الإعدادات', 'الإدارة'),
        ('view_reports', 'عرض التقارير', 'التقارير'),
        ('manage_permissions', 'إدارة الصلاحيات', 'الإدارة'),
        ('view_all_data', 'عرض جميع البيانات', 'البيانات'),
        ('manage_dynamic_fields', 'إدارة الحقول الديناميكية', 'الإدارة'),
        ('view_materials', 'عرض المواد', 'المواد'),
        ('add_materials', 'إضافة المواد', 'المواد'),
        ('edit_materials', 'تعديل المواد', 'المواد'),
        ('delete_materials', 'حذف المواد', 'المواد'),
        ('distribute_materials', 'توزيع المواد', 'المواد'),
        ('view_material_distributions', 'عرض سجل توزيع المواد', 'المواد'),
        ('manage_material_distributions', 'إدارة توزيع المواد', 'المواد'),
        ('reset_citizens_data', 'تصفير بيانات المواطنين', 'تصفير قاعدة البيانات'),
        ('reset_users_data', 'تصفير بيانات المستخدمين', 'تصفير قاعدة البيانات'),
        ('reset_materials_data', 'تصفير بيانات المواد', 'تصفير قاعدة البيانات'),
        ('reset_all_data', 'تصفير جميع البيانات (عدا الإدمن)', 'تصفير قاعدة البيانات'),
        ('import_citizens_excel', 'استيراد بيانات المواطنين من Excel', 'استيراد البيانات'),
        ('manage_telegram_backup', 'إدارة النسخ الاحتياطي التلقائي عبر تيليجرام', 'النسخ الاحتياطية')
    ]
    
    for perm_name, perm_desc, perm_category in default_permissions:
        c.execute("SELECT * FROM permissions WHERE name = ?", (perm_name,))
        if not c.fetchone():
            c.execute("INSERT INTO permissions (name, description, category) VALUES (?, ?, ?)",
                     (perm_name, perm_desc, perm_category))
    
    conn.commit()
    conn.close()

# Check if site is active
def is_site_active():
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("SELECT site_status FROM settings ORDER BY id DESC LIMIT 1")
    result = c.fetchone()
    conn.close()
    return result[0] == 'active' if result else True

# Permission checking functions
def has_permission(user_id, permission_name):
    """التحقق من صلاحية المستخدم"""
    conn = get_db_connection()
    c = conn.cursor()
    
    # التحقق من المدير الرئيسي
    c.execute("SELECT is_admin FROM users WHERE id = ?", (user_id,))
    user = c.fetchone()
    if user and user[0] == 1:  # مدير رئيسي له جميع الصلاحيات
        conn.close()
        return True
    
    # التحقق من الصلاحية المحددة
    c.execute("""
        SELECT 1 FROM user_permissions up 
        JOIN permissions p ON up.permission_id = p.id 
        WHERE up.user_id = ? AND p.name = ?
    """, (user_id, permission_name))
    result = c.fetchone()
    conn.close()
    return result is not None

def get_user_permissions(user_id):
    """الحصول على جميع صلاحيات المستخدم"""
    conn = get_db_connection()
    c = conn.cursor()
    
    # التحقق من المدير الرئيسي
    c.execute("SELECT is_admin FROM users WHERE id = ?", (user_id,))
    user = c.fetchone()
    if user and user[0] == 1:
        c.execute("SELECT name, description, category FROM permissions ORDER BY category, name")
        permissions = c.fetchall()
        conn.close()
        return permissions
    
    # الحصول على صلاحيات المستخدم المحددة
    c.execute("""
        SELECT p.name, p.description, p.category 
        FROM user_permissions up 
        JOIN permissions p ON up.permission_id = p.id 
        WHERE up.user_id = ?
        ORDER BY p.category, p.name
    """, (user_id,))
    permissions = c.fetchall()
    conn.close()
    return permissions

def get_all_permissions():
    """الحصول على جميع الصلاحيات المتاحة"""
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("SELECT id, name, description, category FROM permissions ORDER BY category, name")
    permissions = c.fetchall()
    conn.close()
    return permissions

def assign_default_permissions(user_id, granted_by_user_id=1):
    """تطبيق الصلاحيات الافتراضية للمستخدمين الجدد العاديين"""
    # قائمة الصلاحيات الافتراضية للمستخدمين العاديين
    default_permission_names = [
        'view_citizens',      # عرض بيانات المواطنين
        'add_citizens',       # إضافة بيانات المواطنين  
        'edit_citizens',      # تعديل بيانات المواطنين
        'export_data',        # تصدير البيانات الأساسي
        'view_reports'        # عرض التقارير
    ]
    
    conn = get_db_connection()
    c = conn.cursor()
    
    try:
        # الحصول على معرفات الصلاحيات الافتراضية
        permission_ids = []
        for perm_name in default_permission_names:
            c.execute("SELECT id FROM permissions WHERE name = ?", (perm_name,))
            result = c.fetchone()
            if result:
                permission_ids.append(result[0])
        
        # تطبيق الصلاحيات الافتراضية
        for perm_id in permission_ids:
            try:
                c.execute("""
                    INSERT INTO user_permissions (user_id, permission_id, granted_by)
                    VALUES (?, ?, ?)
                """, (user_id, perm_id, granted_by_user_id))
            except sqlite3.IntegrityError:
                # تجاهل إذا كانت الصلاحية موجودة بالفعل
                pass
        
        conn.commit()
        return True
    
    except Exception as e:
        conn.rollback()
        print(f"خطأ في تطبيق الصلاحيات الافتراضية: {e}")
        return False
    
    finally:
        conn.close()

# إضافة دالة فحص الصلاحيات لقوالب Jinja2
@app.context_processor
def inject_permission_functions():
    """إدراج دوال مساعدة لفحص الصلاحيات في قوالب Jinja2"""
    def check_permission(permission_name):
        """فحص صلاحية المستخدم الحالي"""
        if 'user_id' not in session:
            return False
        return has_permission(session['user_id'], permission_name)
    
    def get_current_user_permissions():
        """الحصول على صلاحيات المستخدم الحالي"""
        if 'user_id' not in session:
            return []
        return get_user_permissions(session['user_id'])
    
    return dict(
        has_permission=check_permission,
        current_user_permissions=get_current_user_permissions
    )

# دوال إدارة الجلسات وإبطالها تلقائياً
def generate_session_token():
    """توليد token فريد للجلسة"""
    return secrets.token_urlsafe(32)

def invalidate_user_session(user_id):
    """إبطال جلسة مستخدم معين عن طريق تحديث session_token"""
    conn = get_db_connection()
    c = conn.cursor()
    new_token = generate_session_token()
    c.execute("UPDATE users SET session_token = ? WHERE id = ?", (new_token, user_id))
    conn.commit()
    conn.close()

def validate_session():
    """التحقق من صحة الجلسة الحالية"""
    if 'user_id' not in session:
        return False
    
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("SELECT session_token FROM users WHERE id = ?", (session['user_id'],))
    result = c.fetchone()
    conn.close()
    
    if not result:
        return False
    
    # إذا لم يكن هناك token في قاعدة البيانات (حساب قديم) والجلسة لا تحتوي على token، اقبل الجلسة
    if result[0] is None and 'session_token' not in session:
        return True
    
    # إذا كان هناك token في قاعدة البيانات، يجب أن يتطابق مع token الجلسة
    if result[0] is not None:
        return result[0] == session.get('session_token')
    
    # حالات أخرى (DB token is None لكن الجلسة تحتوي على token) - رفض
    return False

def check_session_validity():
    """فحص صحة الجلسة وإخراج المستخدم إذا لم تعد صالحة"""
    if 'user_id' in session:
        if not validate_session():
            session.clear()
            flash('تم إنهاء جلستك لأسباب أمنية. يرجى تسجيل الدخول مرة أخرى.', 'warning')
            return redirect(url_for('login'))
    return None

# Telegram Backup Functions
def get_telegram_backup_settings():
    """الحصول على إعدادات النسخ الاحتياطي عبر تيليجرام"""
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("SELECT * FROM telegram_backup_settings ORDER BY id DESC LIMIT 1")
    result = c.fetchone()
    conn.close()
    return result

def send_file_to_telegram(bot_token, chat_id, file_path, caption=None):
    """إرسال ملف إلى تيليجرام"""
    try:
        url = f"https://api.telegram.org/bot{bot_token}/sendDocument"
        
        with open(file_path, 'rb') as file:
            files = {'document': file}
            data = {'chat_id': chat_id}
            if caption:
                data['caption'] = caption
            
            response = requests.post(url, files=files, data=data, timeout=60)
            
        if response.status_code == 200:
            return True, "تم إرسال الملف بنجاح"
        else:
            return False, f"فشل في إرسال الملف: {response.status_code} - {response.text}"
    
    except Exception as e:
        return False, f"خطأ في إرسال الملف: {str(e)}"

def send_backup_to_telegram(backup_type="full", trigger_action="manual"):
    """إنشاء وإرسال نسخة احتياطية إلى تيليجرام"""
    try:
        # الحصول على إعدادات تيليجرام
        settings = get_telegram_backup_settings()
        if not settings or not settings[3]:  # is_enabled
            return False, "النسخ الاحتياطي التلقائي غير مفعل"
        
        if not settings[1] or not settings[2]:  # bot_token, chat_id
            return False, "إعدادات تيليجرام غير مكتملة"
        
        bot_token = settings[1]
        chat_id = settings[2]
        
        # إنشاء النسخة الاحتياطية
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_filename = f"auto_backup_{timestamp}.zip"
        
        # إنشاء نسخة احتياطية شاملة
        backup_buffer = io.BytesIO()
        with zipfile.ZipFile(backup_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # إضافة قاعدة البيانات
            if os.path.exists('database.db'):
                zip_file.write('database.db', 'database.db')
            
            # إضافة ملفات التطبيق
            zip_file.write('app.py', 'app.py')
            
            # إضافة القوالب
            for template_file in os.listdir('templates'):
                if template_file.endswith('.html'):
                    zip_file.write(f'templates/{template_file}', f'templates/{template_file}')
            
            # إضافة الملفات الثابتة إذا كانت موجودة
            if os.path.exists('static'):
                for root, dirs, files in os.walk('static'):
                    for file in files:
                        file_path = os.path.join(root, file)
                        arcname = os.path.relpath(file_path, '.')
                        zip_file.write(file_path, arcname)
            
            # إضافة معلومات النسخة الاحتياطية
            metadata = {
                'backup_date': datetime.now().isoformat(),
                'backup_type': backup_type,
                'trigger_action': trigger_action,
                'system_version': '1.0'
            }
            zip_file.writestr('backup_metadata.json', json.dumps(metadata, ensure_ascii=False, indent=2))
        
        # حفظ النسخة الاحتياطية في ملف مؤقت
        backup_buffer.seek(0)
        with open(backup_filename, 'wb') as f:
            f.write(backup_buffer.getvalue())
        
        # إرسال النسخة الاحتياطية إلى تيليجرام
        caption = f"🔄 نسخة احتياطية تلقائية\n📅 التاريخ: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n⚡ السبب: {trigger_action}\n📁 النوع: {backup_type}"
        
        success, message = send_file_to_telegram(bot_token, chat_id, backup_filename, caption)
        
        # حذف الملف المؤقت
        if os.path.exists(backup_filename):
            os.remove(backup_filename)
        
        if success:
            # تحديث تاريخ آخر نسخة احتياطية
            conn = get_db_connection()
            c = conn.cursor()
            c.execute("UPDATE telegram_backup_settings SET last_backup_sent = CURRENT_TIMESTAMP")
            conn.commit()
            conn.close()
            
            return True, f"تم إرسال النسخة الاحتياطية بنجاح إلى تيليجرام"
        else:
            return False, f"فشل في إرسال النسخة الاحتياطية: {message}"
    
    except Exception as e:
        return False, f"خطأ في إنشاء النسخة الاحتياطية: {str(e)}"

def trigger_automatic_backup(action_type):
    """تفعيل النسخ الاحتياطي التلقائي بناءً على نوع العملية"""
    try:
        # الحصول على إعدادات تيليجرام
        settings = get_telegram_backup_settings()
        if not settings or not settings[3]:  # is_enabled
            return
        
        # فحص ما إذا كان النوع المحدد مفعلاً للنسخ الاحتياطي
        should_backup = False
        
        if action_type == 'citizen' and settings[4]:  # backup_on_citizen_changes
            should_backup = True
        elif action_type == 'user' and settings[5]:  # backup_on_user_changes
            should_backup = True
        elif action_type == 'material' and settings[6]:  # backup_on_material_changes
            should_backup = True
        elif action_type == 'settings' and settings[7]:  # backup_on_settings_changes
            should_backup = True
        elif action_type == 'permission' and settings[8]:  # backup_on_permission_changes
            should_backup = True
        
        if should_backup:
            # تشغيل النسخ الاحتياطي في الخلفية (بدون توقف التطبيق)
            send_backup_to_telegram("auto", f"تعديل في {action_type}")
    
    except Exception as e:
        # تسجيل الخطأ بصمت دون إيقاف التطبيق
        print(f"خطأ في النسخ الاحتياطي التلقائي: {e}")

@app.before_request
def before_request():
    """فحص الجلسة قبل كل طلب"""
    # استثناء صفحات معينة من فحص الجلسة
    excluded_endpoints = ['login', 'static', 'index']
    if request.endpoint and request.endpoint not in excluded_endpoints:
        response = check_session_validity()
        if response:
            return response
    return None

# Routes
@app.route('/', methods=['GET', 'POST'])
def index():
    if not is_site_active() and ('user_id' not in session or not session.get('is_admin')):
        return render_template('maintenance.html')
    
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        return redirect(url_for('login'))
    
    return render_template('login.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    # Allow login even during maintenance for potential admin access
    # The maintenance check will be done after login based on user role
        
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        conn = get_db_connection()
        c = conn.cursor()
        c.execute("SELECT * FROM users WHERE username = ?", (username,))
        user = c.fetchone()
        conn.close()
        
        if user and check_password_hash(user[2], password):
            # إنشاء session token جديد وحفظه في قاعدة البيانات
            session_token = generate_session_token()
            conn = get_db_connection()
            c = conn.cursor()
            c.execute("UPDATE users SET session_token = ? WHERE id = ?", (session_token, user[0]))
            conn.commit()
            conn.close()
            
            # حفظ بيانات الجلسة
            session['user_id'] = user[0]
            session['username'] = user[1]
            session['is_admin'] = user[3]
            session['session_token'] = session_token
            return redirect(url_for('dashboard'))
        else:
            flash('اسم المستخدم أو كلمة المرور غير صحيحة', 'error')
    
    return render_template('login.html')

@app.route('/public_inquiry', methods=['GET', 'POST'])
def public_inquiry():
    """صفحة الاستعلام العام للزوار لمعرفة المواد المسلمة"""
    
    if not is_site_active():
        return render_template('maintenance.html')
    
    citizen_data = None
    materials_received = []
    search_performed = False
    
    if request.method == 'POST':
        search_type = request.form.get('search_type', 'national_id').strip()
        search_term = request.form.get('search_term', '').strip()
        full_name_search = request.form.get('full_name_search', '').strip()
        
        # تحديد نوع البحث والقيم
        search_performed = True
        citizen_result = None
        
        # Validation before opening database connection
        if search_type == 'national_id' and search_term:
            # البحث بالرقم الوطني
            # Server-side validation: التأكد من أن المدخل هو رقم وطني صحيح (11 رقم)
            if not search_term.isdigit() or len(search_term) != 11:
                flash('يجب إدخال رقم وطني صحيح مكون من 11 رقم فقط', 'error')
                return render_template('public_inquiry.html', 
                                     citizen_data=None,
                                     materials_received=[],
                                     search_performed=True)
        elif search_type == 'full_name' and full_name_search:
            # البحث بالاسم الثلاثي
            # Server-side validation: التأكد من أن الاسم ليس فارغاً ولا يحتوي على أرقام فقط
            if len(full_name_search.strip()) < 3:
                flash('يجب إدخال الاسم الثلاثي كاملاً (3 أحرف على الأقل)', 'error')
                return render_template('public_inquiry.html', 
                                     citizen_data=None,
                                     materials_received=[],
                                     search_performed=True)
            
            if full_name_search.strip().isdigit():
                flash('الاسم لا يمكن أن يكون أرقام فقط', 'error')
                return render_template('public_inquiry.html', 
                                     citizen_data=None,
                                     materials_received=[],
                                     search_performed=True)
        else:
            # لم يتم إدخال قيم البحث
            flash('يجب إدخال قيمة للبحث', 'error')
            return render_template('public_inquiry.html', 
                                 citizen_data=None,
                                 materials_received=[],
                                 search_performed=True)
        
        # Open database connection after validation
        conn = get_db_connection()
        c = conn.cursor()
        
        try:
            if search_type == 'national_id' and search_term:
                # البحث بالرقم الوطني (بحث دقيق)
                c.execute("""
                    SELECT id, full_name, status 
                    FROM citizens 
                    WHERE national_id = ?
                """, (search_term,))
                
                citizen_result = c.fetchone()
                
            elif search_type == 'full_name' and full_name_search:
                # البحث بالاسم الثلاثي (بحث دقيق - مطابقة تامة)
                c.execute("""
                    SELECT id, full_name, status 
                    FROM citizens 
                    WHERE full_name = ?
                """, (full_name_search.strip(),))
                
                citizen_result = c.fetchone()
            
            # معالجة نتائج البحث (مشتركة لكلا النوعين)
            if citizen_result:
                citizen_data = {
                    'id': citizen_result[0],
                    'full_name': citizen_result[1],
                    'status': citizen_result[2]
                }
                
                # جلب المواد المسلمة لهذا المواطن (بدون ملاحظات لأمان أكبر)
                c.execute("""
                    SELECT m.name, md.quantity, m.unit, md.distribution_date
                    FROM material_distributions md
                    JOIN materials m ON md.material_id = m.id
                    WHERE md.citizen_id = ?
                    ORDER BY md.distribution_date DESC
                """, (citizen_result[0],))
                
                materials_received = c.fetchall()
                
        finally:
            conn.close()
    
    return render_template('public_inquiry.html', 
                         citizen_data=citizen_data,
                         materials_received=materials_received,
                         search_performed=search_performed)

@app.route('/dashboard')
def dashboard():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Check maintenance mode for non-admin users
    if not is_site_active() and not session.get('is_admin'):
        return render_template('maintenance.html')
    
    conn = get_db_connection()
    c = conn.cursor()
    
    # Get statistics based on user role
    if session.get('is_admin'):
        c.execute("SELECT COUNT(*) FROM citizens")
        total_citizens = c.fetchone()[0]
        
        c.execute("SELECT status, COUNT(*) FROM citizens GROUP BY status")
        status_stats = c.fetchall()
        
        c.execute("SELECT COUNT(*) FROM users")
        total_users = c.fetchone()[0]
    else:
        c.execute("SELECT COUNT(*) FROM citizens WHERE added_by = ?", (session['username'],))
        total_citizens = c.fetchone()[0]
        
        c.execute("SELECT status, COUNT(*) FROM citizens WHERE added_by = ? GROUP BY status", (session['username'],))
        status_stats = c.fetchall()
        
        total_users = 0  # المستخدمون العاديون لا يحتاجون رؤية عدد المستخدمين
    
    conn.close()
    
    return render_template('dashboard.html', 
                         total_citizens=total_citizens,
                         status_stats=status_stats,
                         total_users=total_users)

@app.route('/add_citizen', methods=['GET', 'POST'])
def add_citizen():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        full_name = request.form['full_name']
        national_id = request.form['national_id']
        phone = request.form['phone']
        status = request.form['status']
        family_members = int(request.form['family_members'])
        address = request.form['address']
        notes = request.form.get('notes', '')
        
        # إنشاء كائن البيانات للاحتفاظ بها في حالة الخطأ
        form_data = {
            'full_name': full_name,
            'national_id': national_id,
            'phone': phone,
            'status': status,
            'family_members': family_members,
            'address': address,
            'notes': notes
        }
        
        # Validation
        if len(national_id) != 11 or not national_id.isdigit():
            flash('الرقم الوطني يجب أن يكون 11 رقم', 'error')
            return render_template('add_citizen.html', form_data=form_data)
        
        if len(phone) != 10 or not phone.startswith('09') or not phone.isdigit():
            flash('رقم الجوال يجب أن يكون 10 أرقام ويبدأ ب 09', 'error')
            return render_template('add_citizen.html', form_data=form_data)
        
        if family_members > 50:
            flash('عدد أفراد الأسرة لا يمكن أن يتجاوز 50', 'error')
            return render_template('add_citizen.html', form_data=form_data)
        
        conn = get_db_connection()
        c = conn.cursor()
        
        # Check if national_id already exists
        c.execute("SELECT * FROM citizens WHERE national_id = ?", (national_id,))
        if c.fetchone():
            flash('الرقم الوطني مسجل مسبقاً', 'error')
            conn.close()
            # إعادة عرض الصفحة مع البيانات المدخلة مسبقاً
            form_data = {
                'full_name': full_name,
                'national_id': national_id,
                'phone': phone,
                'status': status,
                'family_members': family_members,
                'address': address,
                'notes': notes
            }
            return render_template('add_citizen.html', form_data=form_data)
        
        try:
            c.execute('''INSERT INTO citizens 
                        (full_name, national_id, phone, status, family_members, address, notes, added_by)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
                     (full_name, national_id, phone, status, family_members, address, notes, session['username']))
            conn.commit()
            flash('تم إضافة البيانات بنجاح', 'success')
            conn.close()
            
            # تفعيل النسخ الاحتياطي التلقائي
            trigger_automatic_backup('citizen')
            
            return redirect(url_for('dashboard'))
        except Exception as e:
            flash('خطأ في إضافة البيانات', 'error')
            conn.close()
    
    return render_template('add_citizen.html')

@app.route('/view_citizens')
def view_citizens():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    page = request.args.get('page', 1, type=int)
    per_page = 20
    search = request.args.get('search', '')
    status_filter = request.args.get('status', '')
    material_filter = request.args.get('material_filter', '')  # received, not_received, multiple
    
    conn = get_db_connection()
    c = conn.cursor()
    
    # Base query with material distribution counts
    if material_filter:
        # Use a subquery to include material distribution counts
        query = """
            SELECT c.*, 
                   COALESCE(md_count.distribution_count, 0) as distribution_count
            FROM citizens c
            LEFT JOIN (
                SELECT citizen_id, COUNT(*) as distribution_count
                FROM material_distributions
                GROUP BY citizen_id
            ) md_count ON c.id = md_count.citizen_id
            WHERE 1=1
        """
    else:
        query = "SELECT * FROM citizens WHERE 1=1"
    
    params = []
    
    # إذا لم يكن لديه صلاحية عرض جميع البيانات، عرض البيانات المضافة من المستخدم فقط
    if not (session.get('is_admin') or has_permission(session['user_id'], 'view_all_data')):
        query += " AND added_by = ?" if 'c.added_by' in query else " AND added_by = ?"
        params.append(session['username'])
    
    if search:
        if 'c.full_name' in query:
            query += " AND (c.full_name LIKE ? OR c.national_id LIKE ? OR c.phone LIKE ?)"
        else:
            query += " AND (full_name LIKE ? OR national_id LIKE ? OR phone LIKE ?)"
        params.extend([f'%{search}%', f'%{search}%', f'%{search}%'])
    
    if status_filter:
        if 'c.status' in query:
            query += " AND c.status = ?"
        else:
            query += " AND status = ?"
        params.append(status_filter)
    
    # Material distribution filtering
    if material_filter == 'received':
        query += " AND COALESCE(md_count.distribution_count, 0) > 0"
    elif material_filter == 'not_received':
        query += " AND COALESCE(md_count.distribution_count, 0) = 0"
    elif material_filter == 'multiple':
        query += " AND COALESCE(md_count.distribution_count, 0) > 1"
    
    if 'c.created_at' in query:
        query += " ORDER BY c.created_at DESC"
    else:
        query += " ORDER BY created_at DESC"
    
    c.execute(query, params)
    all_citizens = c.fetchall()
    
    # Pagination
    offset = (page - 1) * per_page
    citizens = all_citizens[offset:offset + per_page]
    total_pages = (len(all_citizens) + per_page - 1) // per_page
    
    conn.close()
    
    return render_template('view_citizens.html', 
                         citizens=citizens,
                         page=page,
                         total_pages=total_pages,
                         search=search,
                         status_filter=status_filter,
                         material_filter=material_filter)

@app.route('/admin')
def admin():
    if 'user_id' not in session:
        flash('يجب تسجيل الدخول أولاً', 'error')
        return redirect(url_for('login'))
    
    # فحص صلاحيات الإدارة - المدراء أو المستخدمون الذين لديهم صلاحيات إدارية
    user_has_admin_permission = (
        session.get('is_admin') or 
        has_permission(session['user_id'], 'manage_users') or
        has_permission(session['user_id'], 'manage_settings') or
        has_permission(session['user_id'], 'manage_permissions')
    )
    
    if not user_has_admin_permission:
        flash('غير مسموح لك بالوصول لهذه الصفحة', 'error')
        return redirect(url_for('dashboard'))
    
    conn = get_db_connection()
    c = conn.cursor()
    
    # Get all users
    c.execute("SELECT id, username, is_admin, created_at FROM users")
    users = c.fetchall()
    
    # Get settings
    c.execute("SELECT * FROM settings ORDER BY id DESC LIMIT 1")
    settings = c.fetchone()
    
    conn.close()
    
    return render_template('admin.html', users=users, settings=settings)

@app.route('/telegram_backup_settings', methods=['GET', 'POST'])
@require_permission('manage_telegram_backup')
@csrf_protect
def telegram_backup_settings():
    if request.method == 'POST':
        bot_token = request.form.get('bot_token', '').strip()
        chat_id = request.form.get('chat_id', '').strip()
        is_enabled = 1 if 'is_enabled' in request.form else 0
        backup_on_citizen_changes = 1 if 'backup_on_citizen_changes' in request.form else 0
        backup_on_user_changes = 1 if 'backup_on_user_changes' in request.form else 0
        backup_on_material_changes = 1 if 'backup_on_material_changes' in request.form else 0
        backup_on_settings_changes = 1 if 'backup_on_settings_changes' in request.form else 0
        backup_on_permission_changes = 1 if 'backup_on_permission_changes' in request.form else 0
        backup_file_name_pattern = request.form.get('backup_file_name_pattern', 'backup_{timestamp}').strip()
        
        # Test connection if enabled and credentials provided
        if is_enabled and bot_token and chat_id:
            # Only test if bot_token is not the masked placeholder
            if bot_token != '***':
                test_success, test_message = test_telegram_connection(bot_token, chat_id)
                if not test_success:
                    flash(f'فشل في اختبار الاتصال مع تيليجرام: {test_message}', 'error')
                conn = get_db_connection()
                c = conn.cursor()
                c.execute("SELECT * FROM telegram_backup_settings ORDER BY id DESC LIMIT 1")
                settings = c.fetchone()
                conn.close()
                return render_template('telegram_backup_settings.html', settings=settings)
        
        conn = get_db_connection()
        c = conn.cursor()
        
        # Update or insert settings
        c.execute("SELECT id, bot_token FROM telegram_backup_settings ORDER BY id DESC LIMIT 1")
        existing = c.fetchone()
        
        # If bot_token is masked, preserve the existing one
        if bot_token == '***' and existing and existing[1]:
            bot_token = existing[1]
        
        if existing:
            c.execute("""UPDATE telegram_backup_settings SET 
                         bot_token = ?, chat_id = ?, is_enabled = ?, 
                         backup_on_citizen_changes = ?, backup_on_user_changes = ?, 
                         backup_on_material_changes = ?, backup_on_settings_changes = ?, 
                         backup_on_permission_changes = ?, backup_file_name_pattern = ?,
                         updated_at = CURRENT_TIMESTAMP WHERE id = ?""",
                     (bot_token, chat_id, is_enabled, backup_on_citizen_changes, 
                      backup_on_user_changes, backup_on_material_changes, 
                      backup_on_settings_changes, backup_on_permission_changes,
                      backup_file_name_pattern, existing[0]))
        else:
            c.execute("""INSERT INTO telegram_backup_settings 
                         (bot_token, chat_id, is_enabled, backup_on_citizen_changes, 
                          backup_on_user_changes, backup_on_material_changes, 
                          backup_on_settings_changes, backup_on_permission_changes, 
                          backup_file_name_pattern) 
                         VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                     (bot_token, chat_id, is_enabled, backup_on_citizen_changes, 
                      backup_on_user_changes, backup_on_material_changes, 
                      backup_on_settings_changes, backup_on_permission_changes,
                      backup_file_name_pattern))
        
        conn.commit()
        conn.close()
        
        flash('تم حفظ إعدادات النسخ الاحتياطي التلقائي بنجاح', 'success')
        return redirect(url_for('telegram_backup_settings'))
    
    # GET request
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("SELECT * FROM telegram_backup_settings ORDER BY id DESC LIMIT 1")
    settings = c.fetchone()
    conn.close()
    
    return render_template('telegram_backup_settings.html', settings=settings)

def test_telegram_connection(bot_token, chat_id):
    """اختبار الاتصال مع تيليجرام"""
    try:
        url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
        data = {
            'chat_id': chat_id,
            'text': '🔗 اختبار الاتصال مع نظام النسخ الاحتياطي التلقائي\n✅ تم تكوين الإعدادات بنجاح'
        }
        response = requests.post(url, data=data, timeout=10)
        
        if response.status_code == 200:
            return True, "تم الاتصال بنجاح"
        else:
            return False, f"رمز الخطأ: {response.status_code} - {response.text}"
    
    except Exception as e:
        return False, f"خطأ في الاتصال: {str(e)}"

@app.route('/test_telegram_backup', methods=['POST'])
@require_permission('manage_telegram_backup')
@csrf_protect
def test_telegram_backup():
    """اختبار إرسال نسخة احتياطية يدوياً"""
    try:
        success, message = send_backup_to_telegram("manual_test", "اختبار يدوي من لوحة الإدارة")
        
        if success:
            return jsonify({'success': True, 'message': message})
        else:
            return jsonify({'success': False, 'message': message})
    
    except Exception as e:
        return jsonify({'success': False, 'message': f'خطأ في الاختبار: {str(e)}'})

@app.route('/edit_citizen/<int:citizen_id>', methods=['GET', 'POST'])
def edit_citizen(citizen_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    c = conn.cursor()
    
    # التحقق من صلاحية التعديل
    if session.get('is_admin'):
        c.execute("SELECT * FROM citizens WHERE id = ?", (citizen_id,))
    else:
        c.execute("SELECT * FROM citizens WHERE id = ? AND added_by = ?", (citizen_id, session['username']))
    
    citizen = c.fetchone()
    if not citizen:
        flash('غير مسموح لك بتعديل هذا السجل', 'error')
        conn.close()
        return redirect(url_for('view_citizens'))
    
    if request.method == 'POST':
        full_name = request.form['full_name']
        national_id = request.form['national_id']
        phone = request.form['phone']
        status = request.form['status']
        family_members = int(request.form['family_members'])
        address = request.form['address']
        notes = request.form.get('notes', '')
        
        # معالجة assigned_to للمديرين فقط مع validation
        assigned_to = None
        if session.get('is_admin'):
            assigned_to_input = request.form.get('assigned_to', '').strip()
            if assigned_to_input:
                # التحقق من وجود المستخدم
                c.execute("SELECT username FROM users WHERE username = ?", (assigned_to_input,))
                if c.fetchone():
                    assigned_to = assigned_to_input
                else:
                    return render_error('المستخدم المحدد غير موجود')
        
        # Helper function لإعادة عرض الصفحة مع المستخدمين في حالة الخطأ
        def render_error(message):
            flash(message, 'error')
            users = []
            if session.get('is_admin'):
                c.execute("SELECT id, username FROM users ORDER BY username")
                users = c.fetchall()
            conn.close()
            return render_template('edit_citizen.html', citizen=citizen, users=users)
        
        # Validation
        if len(national_id) != 11 or not national_id.isdigit():
            return render_error('الرقم الوطني يجب أن يكون 11 رقم')
        
        if len(phone) != 10 or not phone.startswith('09') or not phone.isdigit():
            return render_error('رقم الجوال يجب أن يكون 10 أرقام ويبدأ ب 09')
        
        if family_members > 50:
            return render_error('عدد أفراد الأسرة لا يمكن أن يتجاوز 50')
        
        # Check if national_id already exists for another citizen
        c.execute("SELECT * FROM citizens WHERE national_id = ? AND id != ?", (national_id, citizen_id))
        if c.fetchone():
            return render_error('الرقم الوطني مسجل مسبقاً لمواطن آخر')
        
        try:
            if session.get('is_admin'):
                # المديرون يمكنهم تحديث assigned_to
                # عند نسب البيانات لمستخدم، يصبح ذلك المستخدم هو من أضاف البيانات
                if assigned_to:
                    # نسب البيانات: تحديث كل من assigned_to و added_by
                    c.execute('''UPDATE citizens SET 
                                full_name = ?, national_id = ?, phone = ?, status = ?, 
                                family_members = ?, address = ?, notes = ?, assigned_to = ?, added_by = ?
                                WHERE id = ?''',
                             (full_name, national_id, phone, status, family_members, address, notes, assigned_to, assigned_to, citizen_id))
                else:
                    # إلغاء النسب: تحديث assigned_to فقط (نحافظ على added_by الأصلي)
                    c.execute('''UPDATE citizens SET 
                                full_name = ?, national_id = ?, phone = ?, status = ?, 
                                family_members = ?, address = ?, notes = ?, assigned_to = ?
                                WHERE id = ?''',
                             (full_name, national_id, phone, status, family_members, address, notes, assigned_to, citizen_id))
            else:
                # المستخدمون العاديون لا يمكنهم تحديث assigned_to (نحافظ على القيمة الحالية)
                c.execute('''UPDATE citizens SET 
                            full_name = ?, national_id = ?, phone = ?, status = ?, 
                            family_members = ?, address = ?, notes = ?
                            WHERE id = ?''',
                         (full_name, national_id, phone, status, family_members, address, notes, citizen_id))
            conn.commit()
            flash('تم تحديث البيانات بنجاح', 'success')
            conn.close()
            
            # تفعيل النسخ الاحتياطي التلقائي
            trigger_automatic_backup('citizen')
            
            return redirect(url_for('view_citizens'))
        except Exception as e:
            flash('خطأ في تحديث البيانات', 'error')
            conn.close()
    
    # جلب قائمة المستخدمين للمديرين فقط
    users = []
    if session.get('is_admin'):
        c.execute("SELECT id, username FROM users ORDER BY username")
        users = c.fetchall()
    
    conn.close()
    return render_template('edit_citizen.html', citizen=citizen, users=users)

@app.route('/export')
def export():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Get all users if admin or has manage_users permission
    users = []
    if session.get('is_admin') or has_permission(session['user_id'], 'manage_users'):
        conn = get_db_connection()
        c = conn.cursor()
        c.execute("SELECT username FROM users ORDER BY username")
        users = [row[0] for row in c.fetchall()]
        conn.close()
    
    return render_template('export.html', users=users)

@app.route('/export_advanced', methods=['GET', 'POST'])
def export_advanced():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Get filter parameters
    status_filters = request.values.getlist('status')
    date_from = request.values.get('date_from')
    date_to = request.values.get('date_to')
    search_text = request.values.get('search_text')
    export_format = request.values.get('format', 'excel')
    grouping = request.values.get('grouping', 'single')
    fields = request.values.getlist('fields')
    all_status = request.values.get('all_status')
    all_fields = request.values.get('all_fields')
    preview = request.values.get('preview')
    user_filters = request.values.getlist('users')
    all_users = request.values.get('all_users')
    material_status = request.values.get('material_status', 'all')
    
    # Build query based on material filter
    if material_status != 'all':
        # Use subquery to include material distribution counts
        query = """
            SELECT c.*, 
                   COALESCE(md_count.distribution_count, 0) as distribution_count
            FROM citizens c
            LEFT JOIN (
                SELECT citizen_id, COUNT(*) as distribution_count
                FROM material_distributions
                GROUP BY citizen_id
            ) md_count ON c.id = md_count.citizen_id
            WHERE 1=1
        """
    else:
        query = "SELECT * FROM citizens WHERE 1=1"
    
    params = []
    
    # إذا لم يكن مدير أو لا يملك صلاحية إدارة المستخدمين، عرض البيانات المضافة من المستخدم فقط
    if not (session.get('is_admin') or has_permission(session['user_id'], 'manage_users')):
        if 'c.added_by' in query:
            query += " AND c.added_by = ?"
        else:
            query += " AND added_by = ?"
        params.append(session['username'])
    else:
        # للمدير أو من لديه صلاحية إدارة المستخدمين: تطبيق فلتر المستخدمين
        if not all_users and user_filters:
            placeholders = ','.join(['?' for _ in user_filters])
            if 'c.added_by' in query:
                query += f" AND c.added_by IN ({placeholders})"
            else:
                query += f" AND added_by IN ({placeholders})"
            params.extend(user_filters)
    
    # Apply filters
    if not all_status and status_filters:
        placeholders = ','.join(['?' for _ in status_filters])
        if 'c.status' in query:
            query += f" AND c.status IN ({placeholders})"
        else:
            query += f" AND status IN ({placeholders})"
        params.extend(status_filters)
    
    if date_from:
        if 'c.created_at' in query:
            query += " AND date(c.created_at) >= ?"
        else:
            query += " AND date(created_at) >= ?"
        params.append(date_from)
    
    if date_to:
        if 'c.created_at' in query:
            query += " AND date(c.created_at) <= ?"
        else:
            query += " AND date(created_at) <= ?"
        params.append(date_to)
    
    if search_text:
        if 'c.full_name' in query:
            query += " AND (c.full_name LIKE ? OR c.national_id LIKE ?)"
        else:
            query += " AND (full_name LIKE ? OR national_id LIKE ?)"
        params.extend([f'%{search_text}%', f'%{search_text}%'])
    
    # Apply material distribution filtering
    if material_status == 'received':
        query += " AND COALESCE(md_count.distribution_count, 0) > 0"
    elif material_status == 'not_received':
        query += " AND COALESCE(md_count.distribution_count, 0) = 0"
    elif material_status == 'multiple':
        query += " AND COALESCE(md_count.distribution_count, 0) > 1"
    
    if 'c.created_at' in query:
        query += " ORDER BY c.created_at DESC"
    else:
        query += " ORDER BY created_at DESC"
    
    conn = get_db_connection()
    c = conn.cursor()
    c.execute(query, params)
    citizens = c.fetchall()
    conn.close()
    
    # If preview requested, return JSON
    if preview:
        preview_html = ""
        if len(citizens) > 0:
            preview_html = "<div class='table-responsive'><table class='table table-sm'>"
            preview_html += "<thead><tr><th>الاسم</th><th>الرقم الوطني</th><th>الحالة</th></tr></thead><tbody>"
            for citizen in citizens[:5]:  # Show first 5 records
                preview_html += f"<tr><td>{citizen[1]}</td><td>{citizen[2]}</td><td>{citizen[4]}</td></tr>"
            if len(citizens) > 5:
                preview_html += f"<tr><td colspan='3'>... و {len(citizens) - 5} سجل آخر</td></tr>"
            preview_html += "</tbody></table></div>"
        
        return jsonify({
            'count': len(citizens),
            'preview': preview_html
        })
    
    if not citizens:
        flash('لا توجد بيانات مطابقة للفلاتر المحددة', 'warning')
        return redirect(url_for('export'))
    
    # Determine fields to export
    field_mapping = {
        'full_name': 'الاسم الثلاثي',
        'national_id': 'الرقم الوطني', 
        'phone': 'الجوال',
        'status': 'الحالة',
        'family_members': 'أفراد الأسرة',
        'address': 'العنوان'
    }
    
    if all_fields or not fields:
        selected_fields = list(field_mapping.keys())
    else:
        selected_fields = fields
    
    # Generate filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    if grouping == 'separate' and status_filters and not all_status:
        # Export separate files for each status
        return export_separate_files(citizens, status_filters, export_format, selected_fields, field_mapping, timestamp)
    else:
        # Export single file
        return export_single_file(citizens, export_format, selected_fields, field_mapping, timestamp)

def export_single_file(citizens, export_format, selected_fields, field_mapping, timestamp):
    if export_format == 'excel':
        return export_excel_file(citizens, selected_fields, field_mapping, timestamp)
    elif export_format == 'pdf':
        return export_pdf_file(citizens, selected_fields, field_mapping, timestamp)
    elif export_format == 'csv':
        return export_csv_file(citizens, selected_fields, field_mapping, timestamp)

def export_separate_files(citizens, status_filters, export_format, selected_fields, field_mapping, timestamp):
    # Group citizens by status
    grouped_citizens = {}
    for citizen in citizens:
        status = citizen[4]
        if status not in grouped_citizens:
            grouped_citizens[status] = []
        grouped_citizens[status].append(citizen)
    
    if export_format == 'excel':
        # Create Excel with multiple sheets
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for status, status_citizens in grouped_citizens.items():
                df_data = prepare_dataframe_data(status_citizens, selected_fields, field_mapping)
                df = pd.DataFrame(df_data[1:], columns=df_data[0])
                df.to_excel(writer, sheet_name=status, index=False)
        
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name=f'citizens_by_status_{timestamp}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    # For PDF and CSV, create a zip file with separate files
    import zipfile
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for status, status_citizens in grouped_citizens.items():
            if export_format == 'pdf':
                file_content = create_pdf_content(status_citizens, selected_fields, field_mapping)
                zip_file.writestr(f'{status}_{timestamp}.pdf', file_content)
            elif export_format == 'csv':
                file_content = create_csv_content(status_citizens, selected_fields, field_mapping)
                zip_file.writestr(f'{status}_{timestamp}.csv', file_content)
    
    zip_buffer.seek(0)
    
    return send_file(
        zip_buffer,
        as_attachment=True,
        download_name=f'citizens_by_status_{timestamp}.zip',
        mimetype='application/zip'
    )

def prepare_dataframe_data(citizens, selected_fields, field_mapping):
    # Column mapping
    columns = [field_mapping[field] for field in selected_fields]
    field_indices = {
        'full_name': 1,
        'national_id': 2,
        'phone': 3,
        'status': 4,
        'family_members': 5,
        'address': 6
    }
    
    data = [columns]
    for citizen in citizens:
        row = []
        for field in selected_fields:
            row.append(citizen[field_indices[field]])
        data.append(row)
    
    return data

def export_excel_file(citizens, selected_fields, field_mapping, timestamp):
    data = prepare_dataframe_data(citizens, selected_fields, field_mapping)
    df = pd.DataFrame(data[1:], columns=data[0])
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='البيانات', index=False)
        
        # Style the worksheet
        worksheet = writer.sheets['البيانات']
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Header styling
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f'citizens_data_{timestamp}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

def export_csv_file(citizens, selected_fields, field_mapping, timestamp):
    import csv
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    data = prepare_dataframe_data(citizens, selected_fields, field_mapping)
    for row in data:
        writer.writerow(row)
    
    csv_content = output.getvalue()
    output.close()
    
    csv_buffer = io.BytesIO()
    csv_buffer.write(csv_content.encode('utf-8-sig'))  # UTF-8 with BOM for Excel compatibility
    csv_buffer.seek(0)
    
    return send_file(
        csv_buffer,
        as_attachment=True,
        download_name=f'citizens_data_{timestamp}.csv',
        mimetype='text/csv'
    )

def create_csv_content(citizens, selected_fields, field_mapping):
    import csv
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    data = prepare_dataframe_data(citizens, selected_fields, field_mapping)
    for row in data:
        writer.writerow(row)
    
    content = output.getvalue()
    output.close()
    return content.encode('utf-8-sig')

def export_pdf_file(citizens, selected_fields, field_mapping, timestamp):
    from reportlab.lib.units import cm
    from reportlab.platypus import Spacer, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import arabic_reshaper
    from bidi.algorithm import get_display
    
    output = io.BytesIO()
    
    # Function to process Arabic text
    def process_arabic_text(text):
        if not text or text == '':
            return ''
        try:
            text = str(text)
            # Process Arabic text for proper display
            reshaped_text = arabic_reshaper.reshape(text)
            bidi_text = get_display(reshaped_text)
            return bidi_text
        except Exception as e:
            # If Arabic processing fails, convert to safe text
            return safe_arabic_to_text(text)
    
    def safe_arabic_to_text(text):
        if not text:
            return ''
        text = str(text)
        # Arabic to readable mapping
        replacements = {
            'ارملة': 'Widow',
            'مطلقة': 'Divorced',
            'اعاقة': 'Disability',
            'كبير بالعمر': 'Elderly',
            'حالة صعبة': 'Difficult Case',
            'الاسم الثلاثي': 'Full Name',
            'الرقم الوطني': 'National ID',
            'الجوال': 'Phone',
            'الحالة': 'Status',
            'أفراد الأسرة': 'Family Members',
            'العنوان': 'Address',
            'الملاحظات': 'Notes'
        }
        
        for arabic, english in replacements.items():
            text = text.replace(arabic, english)
        
        return text
    
    # Configure document
    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm,
        title='Citizens Data Report'
    )
    
    # Process data
    raw_data = prepare_dataframe_data(citizens, selected_fields, field_mapping)
    processed_data = []
    
    for row in raw_data:
        processed_row = []
        for cell in row:
            try:
                # Try Arabic processing first
                processed_cell = process_arabic_text(cell)
                processed_row.append(processed_cell)
            except:
                # Fallback to safe conversion
                processed_row.append(safe_arabic_to_text(cell))
        processed_data.append(processed_row)
    
    # Create table
    table = Table(processed_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E5B88')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]))
    
    # Build document
    story = []
    
    # Add title
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Title'],
        fontName='Helvetica-Bold',
        fontSize=16,
        alignment=1,
        spaceAfter=20
    )
    
    try:
        title_text = process_arabic_text('تقرير بيانات المواطنين')
        story.append(Paragraph(title_text, title_style))
    except:
        story.append(Paragraph('Citizens Data Report', title_style))
    
    story.append(Spacer(1, 12))
    
    # Add date
    date_style = ParagraphStyle(
        'Date',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        alignment=1,
        spaceAfter=20
    )
    
    try:
        date_text = process_arabic_text(f'تاريخ التقرير: {datetime.now().strftime("%Y-%m-%d")}')
        story.append(Paragraph(date_text, date_style))
    except:
        story.append(Paragraph(f'Report Date: {datetime.now().strftime("%Y-%m-%d")}', date_style))
    
    story.append(Spacer(1, 12))
    story.append(table)
    
    # Build PDF
    doc.build(story)
    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name=f'citizens_data_{timestamp}.pdf',
        mimetype='application/pdf'
    )

def create_pdf_content(citizens, selected_fields, field_mapping):
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4)
    
    data = prepare_dataframe_data(citizens, selected_fields, field_mapping)
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    
    doc.build([table])
    content = output.getvalue()
    output.close()
    return content

@app.route('/export_excel')
def export_excel():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    
    # تحديد الاستعلام حسب صلاحية المستخدم
    if session.get('is_admin'):
        df = pd.read_sql_query("SELECT * FROM citizens", conn)
    else:
        df = pd.read_sql_query("SELECT * FROM citizens WHERE added_by = ?", conn, params=[session['username']])
    
    conn.close()
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='البيانات', index=False)
    
    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name=f'citizens_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))

# Additional routes
@app.route('/citizen_details/<int:citizen_id>')
def citizen_details(citizen_id):
    if 'user_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("SELECT * FROM citizens WHERE id = ?", (citizen_id,))
    citizen = c.fetchone()
    
    if citizen:
        # Get material distributions for this citizen
        c.execute("""
            SELECT m.name, md.quantity, m.unit, md.distribution_date, u.username, md.notes
            FROM material_distributions md
            JOIN materials m ON md.material_id = m.id
            JOIN users u ON md.distributed_by = u.id
            WHERE md.citizen_id = ?
            ORDER BY md.distribution_date DESC
        """, (citizen_id,))
        distributions = c.fetchall()
        
        conn.close()
        
        return jsonify({
            'full_name': citizen[1],
            'national_id': citizen[2],
            'phone': citizen[3],
            'status': citizen[4],
            'family_members': citizen[5],
            'address': citizen[6],
            'notes': citizen[7],
            'added_by': citizen[8],
            'material_distributions': distributions
        })
    
    conn.close()
    return jsonify({'error': 'Not found'}), 404

@app.route('/delete_citizen/<int:citizen_id>', methods=['DELETE'])
def delete_citizen(citizen_id):
    if 'user_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    if not (session.get('is_admin') or has_permission(session['user_id'], 'delete_citizens')):
        return jsonify({'error': 'غير مسموح لك بحذف البيانات'}), 403
    
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    c.execute("DELETE FROM citizens WHERE id = ?", (citizen_id,))
    conn.commit()
    conn.close()
    
    # تفعيل النسخ الاحتياطي التلقائي
    trigger_automatic_backup('citizen')
    
    return jsonify({'success': True})

@app.route('/add_user', methods=['POST'])
def add_user():
    if 'user_id' not in session or not (session.get('is_admin') or has_permission(session['user_id'], 'manage_users')):
        flash('ليس لديك صلاحية إدارة المستخدمين', 'error')
        return redirect(url_for('dashboard'))
    
    username = request.form['username']
    password = request.form['password']
    is_admin = 1 if 'is_admin' in request.form else 0
    
    # منع المستخدمين غير المديرين من إنشاء حسابات مديرين
    if is_admin == 1 and not session.get('is_admin'):
        flash('ليس لديك صلاحية إنشاء حسابات مديرين', 'error')
        return redirect(url_for('admin'))
    
    conn = get_db_connection()
    c = conn.cursor()
    
    try:
        hashed_password = generate_password_hash(password)
        c.execute("INSERT INTO users (username, password, is_admin) VALUES (?, ?, ?)",
                 (username, hashed_password, is_admin))
        
        # الحصول على معرف المستخدم الجديد
        new_user_id = c.lastrowid
        
        conn.commit()
        conn.close()
        
        # تفعيل النسخ الاحتياطي التلقائي
        trigger_automatic_backup('user')
        
        # تطبيق الصلاحيات الافتراضية للمستخدمين العاديين فقط
        if is_admin == 0:
            success = assign_default_permissions(new_user_id, session['user_id'])
            if success:
                flash('تم إضافة المستخدم بنجاح مع الصلاحيات الافتراضية', 'success')
            else:
                flash('تم إضافة المستخدم بنجاح لكن حدث خطأ في تطبيق بعض الصلاحيات الافتراضية', 'warning')
        else:
            flash('تم إضافة المستخدم المدير بنجاح', 'success')
            
    except sqlite3.IntegrityError:
        flash('اسم المستخدم موجود مسبقاً', 'error')
        conn.close()
    except Exception as e:
        flash(f'خطأ في إضافة المستخدم: {str(e)}', 'error')
        conn.close()
    
    return redirect(url_for('admin'))

@app.route('/apply_default_permissions_all', methods=['POST'])
def apply_default_permissions_all():
    """تطبيق الصلاحيات الافتراضية على جميع المستخدمين العاديين الذين لا يملكون صلاحيات"""
    if 'user_id' not in session or not session.get('is_admin'):
        return redirect(url_for('login'))
    
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    
    try:
        # العثور على المستخدمين العاديين الذين لا يملكون صلاحيات
        c.execute("""
            SELECT u.id, u.username 
            FROM users u
            WHERE u.is_admin = 0 
            AND u.id NOT IN (
                SELECT DISTINCT user_id 
                FROM user_permissions
            )
        """)
        
        users_without_permissions = c.fetchall()
        success_count = 0
        
        conn.close()
        
        # تطبيق الصلاحيات الافتراضية لكل مستخدم
        for user_id, username in users_without_permissions:
            if assign_default_permissions(user_id, session['user_id']):
                success_count += 1
        
        if success_count > 0:
            flash(f'تم تطبيق الصلاحيات الافتراضية على {success_count} مستخدم بنجاح', 'success')
        else:
            flash('جميع المستخدمين العاديين لديهم صلاحيات بالفعل أو حدث خطأ', 'info')
            
    except Exception as e:
        flash(f'خطأ في تطبيق الصلاحيات: {str(e)}', 'error')
        conn.close()
    
    return redirect(url_for('admin'))

@app.route('/delete_user/<int:user_id>', methods=['DELETE'])
def delete_user(user_id):
    if 'user_id' not in session or not (session.get('is_admin') or has_permission(session['user_id'], 'manage_users')):
        return jsonify({'error': 'Unauthorized'}), 401
    
    # إخراج المستخدم من النظام تلقائياً قبل حذفه
    invalidate_user_session(user_id)
    
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("DELETE FROM users WHERE id = ? AND username != 'admin'", (user_id,))
    conn.commit()
    conn.close()
    
    # تفعيل النسخ الاحتياطي التلقائي
    trigger_automatic_backup('user')
    
    return jsonify({'success': True})

@app.route('/update_settings', methods=['POST'])
def update_settings():
    if 'user_id' not in session or not session.get('is_admin'):
        return redirect(url_for('login'))
    
    site_name = request.form['site_name']
    site_status = request.form['site_status']
    
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("UPDATE settings SET site_name = ?, site_status = ?, updated_at = CURRENT_TIMESTAMP",
             (site_name, site_status))
    conn.commit()
    conn.close()
    
    # تفعيل النسخ الاحتياطي التلقائي
    trigger_automatic_backup('settings')
    
    flash('تم تحديث الإعدادات بنجاح', 'success')
    return redirect(url_for('admin'))

@app.route('/backup')
def backup():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    if not (session.get('is_admin') or has_permission(session['user_id'], 'backup_database')):
        flash('غير مسموح لك بعمل نسخ احتياطية', 'error')
        return redirect(url_for('dashboard'))
    
    backup_name = f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
    shutil.copy2('database.db', backup_name)
    
    return send_file(backup_name, as_attachment=True)

@app.route('/backup_full')
def backup_full():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    if not (session.get('is_admin') or has_permission(session['user_id'], 'backup_database')):
        flash('غير مسموح لك بعمل نسخ احتياطية كاملة', 'error')
        return redirect(url_for('dashboard'))
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_filename = f"full_backup_{timestamp}.zip"
    
    # Create comprehensive backup
    backup_buffer = io.BytesIO()
    with zipfile.ZipFile(backup_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        # Add database
        if os.path.exists('database.db'):
            zip_file.write('database.db', 'database.db')
        
        # Add application files
        zip_file.write('app.py', 'app.py')
        
        # Add templates
        for template_file in os.listdir('templates'):
            if template_file.endswith('.html'):
                zip_file.write(f'templates/{template_file}', f'templates/{template_file}')
        
        # Add static files if they exist
        if os.path.exists('static'):
            for root, dirs, files in os.walk('static'):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, '.')
                    zip_file.write(file_path, arcname)
        
        # Add backup metadata
        metadata = {
            'backup_date': datetime.now().isoformat(),
            'backup_type': 'full',
            'admin_user': session['username'],
            'version': '1.0'
        }
        zip_file.writestr('backup_metadata.json', json.dumps(metadata, ensure_ascii=False, indent=2))
    
    backup_buffer.seek(0)
    
    return send_file(
        backup_buffer,
        as_attachment=True,
        download_name=backup_filename,
        mimetype='application/zip'
    )

@app.route('/restore_backup', methods=['GET', 'POST'])
def restore_backup():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    if not (session.get('is_admin') or has_permission(session['user_id'], 'restore_database')):
        flash('غير مسموح لك باستعادة نسخ احتياطية', 'error')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        if 'backup_file' not in request.files:
            flash('لم يتم اختيار ملف', 'error')
            return redirect(url_for('restore_backup'))
        
        file = request.files['backup_file']
        if file.filename == '':
            flash('لم يتم اختيار ملف', 'error')
            return redirect(url_for('restore_backup'))
        
        if file and file.filename and (file.filename.endswith('.db') or file.filename.endswith('.zip')):
            try:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                
                if file.filename.endswith('.db'):
                    # Backup current database first
                    shutil.copy2('database.db', f'database_backup_before_restore_{timestamp}.db')
                    # Restore database
                    file.save('database.db')
                    flash('تم استعادة قاعدة البيانات بنجاح', 'success')
                
                elif file.filename.endswith('.zip'):
                    # Create comprehensive backup of current state
                    current_backup_filename = f'complete_backup_before_restore_{timestamp}.zip'
                    with zipfile.ZipFile(current_backup_filename, 'w', zipfile.ZIP_DEFLATED) as backup_zip:
                        # Backup current database
                        if os.path.exists('database.db'):
                            backup_zip.write('database.db', 'database.db')
                        # Backup current app.py
                        backup_zip.write('app.py', 'app.py')
                        # Backup templates
                        if os.path.exists('templates'):
                            for template_file in os.listdir('templates'):
                                if template_file.endswith('.html'):
                                    backup_zip.write(f'templates/{template_file}', f'templates/{template_file}')
                        # Backup static files
                        if os.path.exists('static'):
                            for root, dirs, files in os.walk('static'):
                                for file_name in files:
                                    file_path = os.path.join(root, file_name)
                                    arcname = os.path.relpath(file_path, '.')
                                    backup_zip.write(file_path, arcname)
                    
                    # Extract and restore from zip
                    temp_path = f'temp_restore_{timestamp}'
                    os.makedirs(temp_path, exist_ok=True)
                    
                    file.save(f'{temp_path}/restore.zip')
                    
                    restored_files = []
                    with zipfile.ZipFile(f'{temp_path}/restore.zip', 'r') as zip_file:
                        zip_file.extractall(temp_path)
                        
                        # Restore database if present
                        if os.path.exists(f'{temp_path}/database.db'):
                            shutil.copy2(f'{temp_path}/database.db', 'database.db')
                            restored_files.append('قاعدة البيانات')
                        
                        # Restore app.py if present (with caution - this could break the running app)
                        # We'll skip app.py restoration to avoid breaking the system
                        # if os.path.exists(f'{temp_path}/app.py'):
                        #     shutil.copy2(f'{temp_path}/app.py', 'app.py')
                        #     restored_files.append('ملف التطبيق الرئيسي')
                        
                        # Restore templates if present
                        templates_path = f'{temp_path}/templates'
                        if os.path.exists(templates_path):
                            if not os.path.exists('templates'):
                                os.makedirs('templates')
                            for template_file in os.listdir(templates_path):
                                if template_file.endswith('.html'):
                                    shutil.copy2(f'{templates_path}/{template_file}', f'templates/{template_file}')
                            restored_files.append('قوالب HTML')
                        
                        # Restore static files if present
                        static_path = f'{temp_path}/static'
                        if os.path.exists(static_path):
                            if not os.path.exists('static'):
                                os.makedirs('static')
                            for root, dirs, files in os.walk(static_path):
                                for file_name in files:
                                    source_file = os.path.join(root, file_name)
                                    relative_path = os.path.relpath(source_file, static_path)
                                    target_file = os.path.join('static', relative_path)
                                    target_dir = os.path.dirname(target_file)
                                    if not os.path.exists(target_dir):
                                        os.makedirs(target_dir)
                                    shutil.copy2(source_file, target_file)
                            restored_files.append('الملفات الثابتة')
                    
                    # Cleanup
                    shutil.rmtree(temp_path)
                    
                    if restored_files:
                        flash(f'تم استعادة النسخة الاحتياطية الكاملة بنجاح. تم استعادة: {", ".join(restored_files)}', 'success')
                    else:
                        flash('تم استخراج النسخة الاحتياطية ولكن لم يتم العثور على ملفات صالحة للاستعادة', 'warning')
                
                return redirect(url_for('admin'))
                
            except Exception as e:
                flash(f'خطأ في استعادة النسخة الاحتياطية: {str(e)}', 'error')
        else:
            flash('نوع الملف غير مدعوم. يجب أن يكون .db أو .zip', 'error')
    
    return render_template('restore_backup.html')

@app.route('/restore_backup_enhanced', methods=['GET', 'POST'])
def restore_backup_enhanced():
    """نظام استعادة النسخة الاحتياطية المحسن مع خيارات متعددة"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    if not (session.get('is_admin') or has_permission(session['user_id'], 'restore_database')):
        flash('غير مسموح لك باستعادة نسخ احتياطية', 'error')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        if 'backup_file' not in request.files:
            flash('لم يتم اختيار ملف', 'error')
            return redirect(url_for('restore_backup_enhanced'))
        
        file = request.files['backup_file']
        restore_type = request.form.get('restore_type', 'full')
        
        if file.filename == '':
            flash('لم يتم اختيار ملف', 'error')
            return redirect(url_for('restore_backup_enhanced'))
        
        if not (file and file.filename and file.filename.endswith('.zip')):
            flash('يجب أن يكون الملف من نوع .zip للاستعادة المحسنة', 'error')
            return redirect(url_for('restore_backup_enhanced'))
        
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            result = perform_enhanced_restore(file, restore_type, timestamp)
            
            if result['success']:
                flash(result['message'], 'success')
            else:
                flash(result['message'], 'error')
                
        except Exception as e:
            flash(f'خطأ في استعادة النسخة الاحتياطية: {str(e)}', 'error')
        
        return redirect(url_for('admin'))
    
    return render_template('restore_backup_enhanced.html')

def perform_enhanced_restore(file, restore_type, timestamp):
    """تنفيذ الاستعادة المحسنة حسب النوع المحدد"""
    try:
        # إنشاء مجلد مؤقت للاستخراج
        temp_path = f'temp_restore_{timestamp}'
        os.makedirs(temp_path, exist_ok=True)
        
        # حفظ واستخراج الملف
        file.save(f'{temp_path}/restore.zip')
        
        restored_components = []
        
        with zipfile.ZipFile(f'{temp_path}/restore.zip', 'r') as zip_file:
            # فحص أمان الملفات قبل الاستخراج
            if not validate_zip_contents(zip_file):
                return {'success': False, 'message': 'ملف ZIP يحتوي على مسارات غير آمنة'}
            
            # استخراج آمن للملفات
            safe_extract_zip(zip_file, temp_path)
            
            # التحقق من وجود ملف البيانات الوصفية
            metadata_path = f'{temp_path}/backup_metadata.json'
            if os.path.exists(metadata_path):
                with open(metadata_path, 'r', encoding='utf-8') as f:
                    metadata = json.loads(f.read())
                    backup_type = metadata.get('backup_type', 'unknown')
                    backup_date = metadata.get('backup_date', 'unknown')
            
            if restore_type == 'structure_only':
                result = restore_structure_only(temp_path, timestamp)
                
            elif restore_type == 'data_only':
                result = restore_data_only(temp_path, timestamp)
                
            elif restore_type == 'full':
                result = restore_full_backup(temp_path, timestamp)
                
            else:
                result = {'success': False, 'message': 'نوع الاستعادة غير مدعوم'}
        
        # تنظيف المجلد المؤقت
        shutil.rmtree(temp_path)
        return result
        
    except Exception as e:
        # تنظيف في حالة الخطأ
        if os.path.exists(temp_path):
            shutil.rmtree(temp_path)
        return {'success': False, 'message': f'خطأ في المعالجة: {str(e)}'}

def restore_structure_only(temp_path, timestamp):
    """استعادة البنية فقط (الكود + القوالب + الملفات الثابتة) مع الحفاظ على البيانات"""
    restored_components = []
    
    try:
        # إنشاء نسخة احتياطية من الحالة الحالية
        create_current_backup_before_restore(timestamp, backup_data=False)
        
        # استعادة ملف التطبيق الرئيسي (app.py)
        app_file_path = f'{temp_path}/app.py'
        if os.path.exists(app_file_path):
            # إنشاء نسخة احتياطية من app.py الحالي
            shutil.copy2('app.py', f'app_backup_before_structure_restore_{timestamp}.py')
            shutil.copy2(app_file_path, 'app.py')
            restored_components.append('ملف التطبيق الرئيسي')
        
        # استعادة القوالب
        templates_path = f'{temp_path}/templates'
        if os.path.exists(templates_path):
            if not os.path.exists('templates'):
                os.makedirs('templates')
            for template_file in os.listdir(templates_path):
                if template_file.endswith('.html'):
                    shutil.copy2(f'{templates_path}/{template_file}', f'templates/{template_file}')
            restored_components.append('قوالب HTML')
        
        # استعادة الملفات الثابتة
        static_path = f'{temp_path}/static'
        if os.path.exists(static_path):
            if not os.path.exists('static'):
                os.makedirs('static')
            for root, dirs, files in os.walk(static_path):
                for file_name in files:
                    source_file = os.path.join(root, file_name)
                    relative_path = os.path.relpath(source_file, static_path)
                    target_file = os.path.join('static', relative_path)
                    target_dir = os.path.dirname(target_file)
                    if not os.path.exists(target_dir):
                        os.makedirs(target_dir)
                    shutil.copy2(source_file, target_file)
            restored_components.append('الملفات الثابتة')
        
        if restored_components:
            message = f'تم استعادة البنية بنجاح مع الحفاظ على البيانات الحالية. تم استعادة: {", ".join(restored_components)}'
            return {'success': True, 'message': message}
        else:
            return {'success': False, 'message': 'لم يتم العثور على ملفات البنية في النسخة الاحتياطية'}
    
    except Exception as e:
        return {'success': False, 'message': f'خطأ في استعادة البنية: {str(e)}'}

def restore_data_only(temp_path, timestamp):
    """استعادة البيانات فقط (قاعدة البيانات)"""
    try:
        database_path = f'{temp_path}/database.db'
        if os.path.exists(database_path):
            # إنشاء نسخة احتياطية من قاعدة البيانات الحالية
            shutil.copy2('database.db', f'database_backup_before_data_restore_{timestamp}.db')
            # استعادة قاعدة البيانات
            shutil.copy2(database_path, 'database.db')
            return {'success': True, 'message': 'تم استعادة البيانات بنجاح مع الحفاظ على البنية الحالية'}
        else:
            return {'success': False, 'message': 'لم يتم العثور على قاعدة البيانات في النسخة الاحتياطية'}
    
    except Exception as e:
        return {'success': False, 'message': f'خطأ في استعادة البيانات: {str(e)}'}

def restore_full_backup(temp_path, timestamp):
    """استعادة البنية والبيانات معاً (استعادة كاملة)"""
    restored_components = []
    
    try:
        # إنشاء نسخة احتياطية شاملة من الحالة الحالية
        create_current_backup_before_restore(timestamp, backup_data=True)
        
        # استعادة قاعدة البيانات
        database_path = f'{temp_path}/database.db'
        if os.path.exists(database_path):
            shutil.copy2(database_path, 'database.db')
            restored_components.append('قاعدة البيانات')
        
        # استعادة ملف التطبيق الرئيسي
        app_file_path = f'{temp_path}/app.py'
        if os.path.exists(app_file_path):
            shutil.copy2(app_file_path, 'app.py')
            restored_components.append('ملف التطبيق الرئيسي')
        
        # استعادة القوالب
        templates_path = f'{temp_path}/templates'
        if os.path.exists(templates_path):
            if not os.path.exists('templates'):
                os.makedirs('templates')
            for template_file in os.listdir(templates_path):
                if template_file.endswith('.html'):
                    shutil.copy2(f'{templates_path}/{template_file}', f'templates/{template_file}')
            restored_components.append('قوالب HTML')
        
        # استعادة الملفات الثابتة
        static_path = f'{temp_path}/static'
        if os.path.exists(static_path):
            if not os.path.exists('static'):
                os.makedirs('static')
            for root, dirs, files in os.walk(static_path):
                for file_name in files:
                    source_file = os.path.join(root, file_name)
                    relative_path = os.path.relpath(source_file, static_path)
                    target_file = os.path.join('static', relative_path)
                    target_dir = os.path.dirname(target_file)
                    if not os.path.exists(target_dir):
                        os.makedirs(target_dir)
                    shutil.copy2(source_file, target_file)
            restored_components.append('الملفات الثابتة')
        
        if restored_components:
            message = f'تم استعادة النسخة الاحتياطية الكاملة بنجاح. تم استعادة: {", ".join(restored_components)}'
            return {'success': True, 'message': message}
        else:
            return {'success': False, 'message': 'لم يتم العثور على ملفات صالحة للاستعادة في النسخة الاحتياطية'}
    
    except Exception as e:
        return {'success': False, 'message': f'خطأ في الاستعادة الكاملة: {str(e)}'}

def create_current_backup_before_restore(timestamp, backup_data=True):
    """إنشاء نسخة احتياطية من الحالة الحالية قبل الاستعادة"""
    try:
        backup_filename = f'auto_backup_before_restore_{timestamp}.zip'
        
        with zipfile.ZipFile(backup_filename, 'w', zipfile.ZIP_DEFLATED) as backup_zip:
            # نسخ احتياطية من قاعدة البيانات إذا كان مطلوباً
            if backup_data and os.path.exists('database.db'):
                backup_zip.write('database.db', 'database.db')
            
            # نسخ احتياطية من app.py
            if os.path.exists('app.py'):
                backup_zip.write('app.py', 'app.py')
            
            # نسخ احتياطية من القوالب
            if os.path.exists('templates'):
                for template_file in os.listdir('templates'):
                    if template_file.endswith('.html'):
                        backup_zip.write(f'templates/{template_file}', f'templates/{template_file}')
            
            # نسخ احتياطية من الملفات الثابتة
            if os.path.exists('static'):
                for root, dirs, files in os.walk('static'):
                    for file_name in files:
                        file_path = os.path.join(root, file_name)
                        arcname = os.path.relpath(file_path, '.')
                        backup_zip.write(file_path, arcname)
            
            # إضافة البيانات الوصفية
            metadata = {
                'backup_date': datetime.now().isoformat(),
                'backup_type': 'auto_before_restore',
                'admin_user': session.get('username', 'unknown'),
                'version': '1.0',
                'includes_data': backup_data
            }
            backup_zip.writestr('backup_metadata.json', json.dumps(metadata, ensure_ascii=False, indent=2))
        
        return True
    
    except Exception as e:
        print(f"خطأ في إنشاء النسخة الاحتياطية التلقائية: {e}")
        return False

def validate_zip_contents(zip_file):
    """التحقق من أن محتويات ملف ZIP آمنة ولا تحتوي على مسارات خطرة"""
    allowed_files = {
        'app.py', 'database.db', 'backup_metadata.json'
    }
    allowed_prefixes = ('templates/', 'static/')
    
    # حدود الحماية من قنابل الضغط
    MAX_FILES = 1000
    MAX_TOTAL_SIZE = 100 * 1024 * 1024  # 100 MB
    MAX_FILE_SIZE = 50 * 1024 * 1024    # 50 MB per file
    MAX_COMPRESSION_RATIO = 100         # نسبة ضغط قصوى
    
    try:
        file_list = zip_file.namelist()
        
        # فحص عدد الملفات
        if len(file_list) > MAX_FILES:
            return False
        
        total_uncompressed_size = 0
        
        for member_name in file_list:
            try:
                info = zip_file.getinfo(member_name)
            except KeyError:
                return False
            
            # تحويل المسار إلى مسار آمن
            member_path = os.path.normpath(member_name)
            
            # رفض المسارات المطلقة
            if os.path.isabs(member_path):
                return False
            
            # رفض المسارات التي تحتوي على ".."
            if '..' in member_path or '../' in member_name or '..\\' in member_name:
                return False
            
            # رفض المسارات التي تبدأ بـ "/" أو "\"
            if member_path.startswith(('/','\\')) or member_name.startswith(('/','\\', '\\\\')):
                return False
            
            # رفض أسماء الملفات التي تحتوي على محارف Windows الخطرة
            if any(char in member_name for char in [':', '*', '?', '"', '<', '>', '|']):
                return False
            
            # التحقق من الروابط المميزة والروابط الصلبة
            if info.external_attr & 0o170000 == 0o120000:  # symlink check
                return False
            
            # فحص حجم الملف المضغوط وغير المضغوط
            if info.file_size > MAX_FILE_SIZE:
                return False
            
            total_uncompressed_size += info.file_size
            if total_uncompressed_size > MAX_TOTAL_SIZE:
                return False
            
            # فحص نسبة الضغط لتجنب قنابل الضغط
            if info.compress_size > 0:
                compression_ratio = info.file_size / info.compress_size
                if compression_ratio > MAX_COMPRESSION_RATIO:
                    return False
            
            # التحقق من أن الملف مسموح به (تخطي المجلدات)
            if not member_name.endswith('/'):
                if member_path not in allowed_files:
                    # التحقق من أن الملف يبدأ بأحد البادئات المسموحة
                    if not any(member_path.startswith(prefix) for prefix in allowed_prefixes):
                        return False
        
        return True
    
    except Exception:
        return False

def safe_extract_zip(zip_file, extract_to):
    """استخراج آمن لملفات ZIP مع حماية من Zip Slip"""
    for member in zip_file.namelist():
        # تحويل المسار إلى مسار آمن
        member_path = os.path.normpath(member)
        
        # التأكد من أن المسار النهائي داخل مجلد الاستخراج
        full_path = os.path.join(extract_to, member_path)
        full_path = os.path.normpath(full_path)
        
        # التحقق من أن المسار لا يخرج من مجلد الاستخراج
        if not full_path.startswith(os.path.normpath(extract_to) + os.sep):
            if not full_path == os.path.normpath(extract_to):
                continue  # تخطي الملفات غير الآمنة
        
        # إنشاء المجلدات اللازمة
        dir_path = os.path.dirname(full_path)
        if dir_path and not os.path.exists(dir_path):
            os.makedirs(dir_path, exist_ok=True)
        
        # استخراج الملف إذا لم يكن مجلداً
        if not member.endswith('/'):
            with zip_file.open(member) as source, open(full_path, 'wb') as target:
                target.write(source.read())

@app.route('/edit_user/<int:user_id>', methods=['GET', 'POST'])
def edit_user(user_id):
    if 'user_id' not in session or not (session.get('is_admin') or has_permission(session['user_id'], 'manage_users')):
        flash('ليس لديك صلاحية إدارة المستخدمين', 'error')
        return redirect(url_for('dashboard'))
    
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    
    # Get user data
    c.execute("SELECT id, username, is_admin FROM users WHERE id = ?", (user_id,))
    user = c.fetchone()
    
    if not user:
        flash('المستخدم غير موجود', 'error')
        conn.close()
        return redirect(url_for('admin'))
    
    # Prevent editing admin user by non-admin or editing self
    if user[1] == 'admin' and session['username'] != 'admin':
        flash('غير مسموح بتعديل المستخدم الرئيسي', 'error')
        conn.close()
        return redirect(url_for('admin'))
    
    # منع المستخدمين غير المديرين من تعديل أي مستخدم مدير
    if user[2] == 1 and not session.get('is_admin'):  # user[2] is is_admin field
        flash('ليس لديك صلاحية تعديل حسابات المديرين', 'error')
        conn.close()
        return redirect(url_for('admin'))
    
    if request.method == 'POST':
        new_username = request.form['username']
        new_password = request.form.get('password')
        is_admin = 1 if 'is_admin' in request.form else 0
        
        # منع المستخدمين من تعديل حساباتهم الشخصية
        if user_id == session['user_id']:
            flash('لا يمكنك تعديل حسابك الشخصي من هذه الصفحة', 'error')
            conn.close()
            return redirect(url_for('admin'))
        
        # منع المستخدمين غير المديرين من ترقية المستخدمين لمديرين
        if is_admin == 1 and not session.get('is_admin'):
            flash('ليس لديك صلاحية ترقية المستخدمين لمديرين', 'error')
            conn.close()
            return redirect(url_for('admin'))
        
        # منع خفض رتبة المدير الأخير المتبقي
        if user[2] == 1 and is_admin == 0:  # محاولة خفض رتبة مدير
            c.execute("SELECT COUNT(*) FROM users WHERE is_admin = 1")
            admin_count = c.fetchone()[0]
            if admin_count <= 1:
                flash('لا يمكن خفض رتبة المدير الأخير المتبقي', 'error')
                conn.close()
                return redirect(url_for('admin'))
        
        # Validation
        if len(new_username.strip()) < 3:
            flash('اسم المستخدم يجب أن يكون 3 أحرف على الأقل', 'error')
            conn.close()
            return render_template('edit_user.html', user=user)
        
        # Check if username exists for another user
        c.execute("SELECT id FROM users WHERE username = ? AND id != ?", (new_username, user_id))
        if c.fetchone():
            flash('اسم المستخدم موجود مسبقاً', 'error')
            conn.close()
            return render_template('edit_user.html', user=user)
        
        try:
            if new_password:
                # Update with new password
                hashed_password = generate_password_hash(new_password)
                c.execute("UPDATE users SET username = ?, password = ?, is_admin = ? WHERE id = ?",
                         (new_username, hashed_password, is_admin, user_id))
            else:
                # Update without changing password
                c.execute("UPDATE users SET username = ?, is_admin = ? WHERE id = ?",
                         (new_username, is_admin, user_id))
            
            conn.commit()
            flash('تم تحديث بيانات المستخدم بنجاح', 'success')
            conn.close()
            
            # تفعيل النسخ الاحتياطي التلقائي
            trigger_automatic_backup('user')
            
            # إخراج المستخدم من النظام تلقائياً بعد تعديل بياناته
            invalidate_user_session(user_id)
            return redirect(url_for('admin'))
            
        except sqlite3.IntegrityError:
            flash('خطأ في تحديث البيانات', 'error')
            conn.close()
    
    conn.close()
    return render_template('edit_user.html', user=user)

@app.route('/manage_permissions/<int:user_id>', methods=['GET', 'POST'])
def manage_permissions(user_id):
    if 'user_id' not in session or not has_permission(session['user_id'], 'manage_permissions'):
        flash('ليس لديك صلاحية إدارة الصلاحيات', 'error')
        return redirect(url_for('dashboard'))
    
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    
    # Get user data
    c.execute("SELECT id, username, is_admin FROM users WHERE id = ?", (user_id,))
    user = c.fetchone()
    
    if not user:
        flash('المستخدم غير موجود', 'error')
        conn.close()
        return redirect(url_for('admin'))
    
    if request.method == 'POST':
        selected_permissions = request.form.getlist('permissions')
        
        # منع المستخدمين من تعديل صلاحياتهم الشخصية
        if user_id == session['user_id']:
            flash('لا يمكنك تعديل صلاحياتك الشخصية', 'error')
            conn.close()
            return redirect(url_for('admin'))
        
        # منع المستخدمين غير المديرين من تعديل صلاحيات المديرين
        if user[2] == 1 and not session.get('is_admin'):  # user[2] is is_admin field
            flash('ليس لديك صلاحية تعديل صلاحيات المديرين', 'error')
            conn.close()
            return redirect(url_for('admin'))
        
        # فحص هرمية الصلاحيات - التأكد من أن المستخدم يمكنه منح الصلاحيات المحددة
        if not session.get('is_admin'):  # المديرون لهم جميع الصلاحيات
            current_user_permissions = get_user_permissions(session['user_id'])
            current_user_permission_ids = []
            
            # الحصول على معرفات صلاحيات المستخدم الحالي
            conn_temp = sqlite3.connect('database.db')
            c_temp = conn_temp.cursor()
            c_temp.execute("""
                SELECT p.id FROM user_permissions up 
                JOIN permissions p ON up.permission_id = p.id 
                WHERE up.user_id = ?
            """, (session['user_id'],))
            current_user_permission_ids = [row[0] for row in c_temp.fetchall()]
            conn_temp.close()
            
            # التحقق من أن جميع الصلاحيات المحددة متاحة للمستخدم الحالي
            for perm_id in selected_permissions:
                if int(perm_id) not in current_user_permission_ids:
                    c.execute("SELECT name FROM permissions WHERE id = ?", (int(perm_id),))
                    perm_name = c.fetchone()
                    perm_name = perm_name[0] if perm_name else "غير معروف"
                    flash(f'لا يمكنك منح صلاحية "{perm_name}" لأنك لا تملكها', 'error')
                    conn.close()
                    return redirect(url_for('admin'))
        
        try:
            # حذف الصلاحيات الحالية للمستخدم
            c.execute("DELETE FROM user_permissions WHERE user_id = ?", (user_id,))
            
            # إضافة الصلاحيات الجديدة
            inserted_count = 0
            for perm_id in selected_permissions:
                try:
                    # التأكد من أن perm_id صحيح
                    perm_id_int = int(perm_id)
                    c.execute("""
                        INSERT INTO user_permissions (user_id, permission_id, granted_by)
                        VALUES (?, ?, ?)
                    """, (user_id, perm_id_int, session['user_id']))
                    inserted_count += 1
                except (ValueError, sqlite3.Error):
                    continue
            
            conn.commit()
            
            # تفعيل النسخ الاحتياطي التلقائي
            trigger_automatic_backup('permission')
            
            if inserted_count > 0:
                flash(f'تم تحديث صلاحيات المستخدم بنجاح. تم حفظ {inserted_count} صلاحية', 'success')
            else:
                flash('تم حذف جميع الصلاحيات من المستخدم', 'info')
            
            conn.close()
            return redirect(url_for('admin'))
            
        except Exception as e:
            flash(f'خطأ في تحديث الصلاحيات: {str(e)}', 'error')
            conn.close()
    
    # الحصول على جميع الصلاحيات
    all_permissions = get_all_permissions()
    
    # الحصول على صلاحيات المستخدم الحالية
    user_permissions = get_user_permissions(user_id)
    user_permissions_list = [perm[0] for perm in user_permissions]  # أسماء الصلاحيات
    
    conn.close()
    return render_template('manage_permissions.html', 
                         user=user, 
                         all_permissions=all_permissions,
                         user_permissions=user_permissions,
                         user_permissions_list=user_permissions_list)

@app.route('/manage_fields')
def manage_fields():
    if 'user_id' not in session or not session.get('is_admin'):
        return redirect(url_for('login'))
    
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    c.execute("SELECT * FROM dynamic_fields ORDER BY field_order, field_name")
    fields = c.fetchall()
    conn.close()
    
    return render_template('manage_fields.html', fields=fields)

@app.route('/add_field', methods=['GET', 'POST'])
def add_field():
    if 'user_id' not in session or not session.get('is_admin'):
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        field_name = request.form['field_name'].strip().lower().replace(' ', '_')
        field_label = request.form['field_label'].strip()
        field_type = request.form['field_type']
        field_options = request.form.get('field_options', '').strip()
        is_required = 1 if 'is_required' in request.form else 0
        validation_rules = request.form.get('validation_rules', '').strip()
        field_order = int(request.form.get('field_order', 0))
        
        # Validation
        if len(field_name) < 2:
            flash('اسم الحقل يجب أن يكون حرفين على الأقل', 'error')
            return render_template('add_field.html')
        
        if len(field_label) < 2:
            flash('تسمية الحقل يجب أن تكون حرفين على الأقل', 'error')
            return render_template('add_field.html')
        
        # Check for reserved field names
        reserved_names = ['id', 'full_name', 'national_id', 'phone', 'status', 'family_members', 'address', 'notes', 'added_by', 'created_at']
        if field_name in reserved_names:
            flash(f'اسم الحقل "{field_name}" محجوز، يرجى اختيار اسم آخر', 'error')
            return render_template('add_field.html')
        
        conn = sqlite3.connect('database.db')
        c = conn.cursor()
        
        try:
            c.execute('''INSERT INTO dynamic_fields 
                        (field_name, field_label, field_type, field_options, is_required, validation_rules, field_order)
                        VALUES (?, ?, ?, ?, ?, ?, ?)''',
                     (field_name, field_label, field_type, field_options, is_required, validation_rules, field_order))
            conn.commit()
            flash('تم إضافة الحقل بنجاح', 'success')
            conn.close()
            return redirect(url_for('manage_fields'))
            
        except sqlite3.IntegrityError:
            flash('اسم الحقل موجود مسبقاً', 'error')
            conn.close()
    
    return render_template('add_field.html')

@app.route('/edit_field/<int:field_id>', methods=['GET', 'POST'])
def edit_field(field_id):
    if 'user_id' not in session or not session.get('is_admin'):
        return redirect(url_for('login'))
    
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    
    c.execute("SELECT * FROM dynamic_fields WHERE id = ?", (field_id,))
    field = c.fetchone()
    
    if not field:
        flash('الحقل غير موجود', 'error')
        conn.close()
        return redirect(url_for('manage_fields'))
    
    if request.method == 'POST':
        field_name = request.form['field_name'].strip().lower().replace(' ', '_')
        field_label = request.form['field_label'].strip()
        field_type = request.form['field_type']
        field_options = request.form.get('field_options', '').strip()
        is_required = 1 if 'is_required' in request.form else 0
        validation_rules = request.form.get('validation_rules', '').strip()
        field_order = int(request.form.get('field_order', 0))
        is_active = 1 if 'is_active' in request.form else 0
        
        # Check if field name exists for another field
        c.execute("SELECT id FROM dynamic_fields WHERE field_name = ? AND id != ?", (field_name, field_id))
        if c.fetchone():
            flash('اسم الحقل موجود مسبقاً لحقل آخر', 'error')
            conn.close()
            return render_template('edit_field.html', field=field)
        
        try:
            c.execute('''UPDATE dynamic_fields SET 
                        field_name = ?, field_label = ?, field_type = ?, field_options = ?,
                        is_required = ?, validation_rules = ?, field_order = ?, is_active = ?
                        WHERE id = ?''',
                     (field_name, field_label, field_type, field_options, is_required, 
                      validation_rules, field_order, is_active, field_id))
            conn.commit()
            flash('تم تحديث الحقل بنجاح', 'success')
            conn.close()
            return redirect(url_for('manage_fields'))
            
        except Exception as e:
            flash('خطأ في تحديث الحقل', 'error')
            conn.close()
    
    conn.close()
    return render_template('edit_field.html', field=field)

@app.route('/delete_field/<int:field_id>', methods=['DELETE'])
def delete_field(field_id):
    if 'user_id' not in session or not session.get('is_admin'):
        return jsonify({'error': 'Unauthorized'}), 401
    
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    
    # Delete field values first
    c.execute("DELETE FROM dynamic_field_values WHERE field_id = ?", (field_id,))
    # Delete the field
    c.execute("DELETE FROM dynamic_fields WHERE id = ?", (field_id,))
    
    conn.commit()
    conn.close()
    
    return jsonify({'success': True})

def get_dynamic_fields():
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    c.execute("SELECT * FROM dynamic_fields WHERE is_active = 1 ORDER BY field_order, field_label")
    fields = c.fetchall()
    conn.close()
    return fields

def get_citizen_dynamic_values(citizen_id):
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    c.execute("""SELECT df.field_name, dfv.field_value 
                FROM dynamic_fields df 
                LEFT JOIN dynamic_field_values dfv ON df.id = dfv.field_id AND dfv.citizen_id = ?
                WHERE df.is_active = 1
                ORDER BY df.field_order, df.field_label""", (citizen_id,))
    values = c.fetchall()
    conn.close()
    return dict(values)

def save_dynamic_field_values(citizen_id, form_data):
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    
    # Get all dynamic fields
    c.execute("SELECT id, field_name FROM dynamic_fields WHERE is_active = 1")
    fields = c.fetchall()
    
    for field_id, field_name in fields:
        field_value = form_data.get(f'dynamic_{field_name}', '')
        
        # Check if value exists
        c.execute("SELECT id FROM dynamic_field_values WHERE citizen_id = ? AND field_id = ?", 
                 (citizen_id, field_id))
        existing = c.fetchone()
        
        if existing:
            # Update existing value
            c.execute("UPDATE dynamic_field_values SET field_value = ? WHERE citizen_id = ? AND field_id = ?",
                     (field_value, citizen_id, field_id))
        else:
            # Insert new value
            c.execute("INSERT INTO dynamic_field_values (citizen_id, field_id, field_value) VALUES (?, ?, ?)",
                     (citizen_id, field_id, field_value))
    
    conn.commit()
    conn.close()

# Material Management Routes
@app.route('/manage_materials')
@require_permission('view_materials')
def manage_materials():
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("""
        SELECT m.id, m.name, m.description, m.unit, m.is_active, u.username, m.created_at
        FROM materials m
        LEFT JOIN users u ON m.created_by = u.id
        ORDER BY m.created_at DESC
    """)
    materials = c.fetchall()
    conn.close()
    
    return render_template('manage_materials.html', materials=materials)

@app.route('/add_material', methods=['GET', 'POST'])
@require_permission('add_materials')
def add_material():
    
    if request.method == 'POST':
        name = request.form['name'].strip()
        description = request.form.get('description', '').strip()
        unit = request.form.get('unit', 'قطعة').strip()
        
        if len(name) < 2:
            flash('اسم المادة يجب أن يكون حرفين على الأقل', 'error')
            return render_template('add_material.html')
        
        conn = get_db_connection()
        c = conn.cursor()
        
        try:
            c.execute("INSERT INTO materials (name, description, unit, created_by) VALUES (?, ?, ?, ?)",
                     (name, description, unit, session['user_id']))
            conn.commit()
            flash('تم إضافة المادة بنجاح', 'success')
            conn.close()
            
            # تفعيل النسخ الاحتياطي التلقائي
            trigger_automatic_backup('material')
            
            return redirect(url_for('manage_materials'))
            
        except sqlite3.IntegrityError:
            flash('اسم المادة موجود مسبقاً', 'error')
            conn.close()
    
    return render_template('add_material.html')

@app.route('/edit_material/<int:material_id>', methods=['GET', 'POST'])
@require_permission('edit_materials')
def edit_material(material_id):
    conn = get_db_connection()
    c = conn.cursor()
    
    c.execute("SELECT * FROM materials WHERE id = ?", (material_id,))
    material = c.fetchone()
    
    if not material:
        flash('المادة غير موجودة', 'error')
        conn.close()
        return redirect(url_for('manage_materials'))
    
    if request.method == 'POST':
        name = request.form['name'].strip()
        description = request.form.get('description', '').strip()
        unit = request.form.get('unit', 'قطعة').strip()
        is_active = 1 if 'is_active' in request.form else 0
        
        if len(name) < 2:
            flash('اسم المادة يجب أن يكون حرفين على الأقل', 'error')
            conn.close()
            return render_template('edit_material.html', material=material)
        
        # Check if name exists for another material
        c.execute("SELECT id FROM materials WHERE name = ? AND id != ?", (name, material_id))
        if c.fetchone():
            flash('اسم المادة موجود مسبقاً لمادة أخرى', 'error')
            conn.close()
            return render_template('edit_material.html', material=material)
        
        try:
            c.execute("UPDATE materials SET name = ?, description = ?, unit = ?, is_active = ? WHERE id = ?",
                     (name, description, unit, is_active, material_id))
            conn.commit()
            flash('تم تحديث المادة بنجاح', 'success')
            conn.close()
            
            # تفعيل النسخ الاحتياطي التلقائي
            trigger_automatic_backup('material')
            
            return redirect(url_for('manage_materials'))
            
        except Exception as e:
            flash('خطأ في تحديث المادة', 'error')
            conn.close()
    
    conn.close()
    return render_template('edit_material.html', material=material)

@app.route('/delete_material/<int:material_id>')
@require_permission('delete_materials')
def delete_material(material_id):
    conn = get_db_connection()
    c = conn.cursor()
    
    # Check if material has distributions
    c.execute("SELECT COUNT(*) FROM material_distributions WHERE material_id = ?", (material_id,))
    count = c.fetchone()[0]
    
    if count > 0:
        flash('لا يمكن حذف المادة لأنها تحتوي على سجلات توزيع', 'error')
    else:
        c.execute("DELETE FROM materials WHERE id = ?", (material_id,))
        conn.commit()
        flash('تم حذف المادة بنجاح', 'success')
        
        # تفعيل النسخ الاحتياطي التلقائي
        trigger_automatic_backup('material')
    
    conn.close()
    return redirect(url_for('manage_materials'))

@app.route('/distribute_material', methods=['GET', 'POST'])
@require_permission('distribute_materials')
def distribute_material():
    conn = get_db_connection()
    c = conn.cursor()
    
    if request.method == 'POST':
        # الحصول على قائمة المواطنين المختارين
        citizen_ids = request.form.getlist('citizen_ids[]')
        material_id = request.form['material_id']
        quantity = int(request.form.get('quantity', 1))
        notes = request.form.get('notes', '').strip()
        
        # التحقق من اختيار مواطن واحد على الأقل
        if not citizen_ids:
            flash('يجب اختيار مواطن واحد على الأقل', 'error')
            conn.close()
            return redirect(url_for('distribute_material'))
        
        try:
            # إنشاء سجل توزيع منفصل لكل مواطن
            for citizen_id in citizen_ids:
                c.execute("""
                    INSERT INTO material_distributions (citizen_id, material_id, quantity, distributed_by, notes)
                    VALUES (?, ?, ?, ?, ?)
                """, (citizen_id, material_id, quantity, session['user_id'], notes))
            
            conn.commit()
            
            # تفعيل النسخ الاحتياطي التلقائي
            trigger_automatic_backup('material')
            
            # رسالة نجاح تتضمن عدد المواطنين
            citizens_count = len(citizen_ids)
            if citizens_count == 1:
                flash('تم تسجيل توزيع المادة للمواطن بنجاح', 'success')
            else:
                flash(f'تم تسجيل توزيع المادة لـ {citizens_count} مواطن بنجاح', 'success')
            
            conn.close()
            return redirect(url_for('distribute_material'))
            
        except Exception as e:
            flash(f'خطأ في تسجيل التوزيع: {str(e)}', 'error')
            conn.close()
    
    # Get active materials
    c.execute("SELECT id, name, unit FROM materials WHERE is_active = 1 ORDER BY name")
    materials = c.fetchall()
    
    # Get citizens
    c.execute("SELECT id, full_name, national_id FROM citizens ORDER BY full_name")
    citizens = c.fetchall()
    
    conn.close()
    
    return render_template('distribute_material.html', materials=materials, citizens=citizens)

@app.route('/material_distributions')
@require_permission('view_material_distributions')
def material_distributions():
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("""
        SELECT md.id, c.full_name, c.national_id, m.name, md.quantity, m.unit,
               u.username, md.distribution_date, md.notes
        FROM material_distributions md
        JOIN citizens c ON md.citizen_id = c.id
        JOIN materials m ON md.material_id = m.id
        JOIN users u ON md.distributed_by = u.id
        ORDER BY md.distribution_date DESC
    """)
    distributions = c.fetchall()
    conn.close()
    
    return render_template('material_distributions.html', distributions=distributions)

@app.route('/export_distributions')
@require_permission('view_material_distributions')
def export_distributions():
    conn = get_db_connection()
    c = conn.cursor()
    
    # Get all material distributions with detailed info
    c.execute("""
        SELECT md.id, c.full_name, c.national_id, c.phone, m.name as material_name, 
               md.quantity, m.unit, u.username, md.distribution_date, md.notes
        FROM material_distributions md
        JOIN citizens c ON md.citizen_id = c.id
        JOIN materials m ON md.material_id = m.id
        JOIN users u ON md.distributed_by = u.id
        ORDER BY md.distribution_date DESC
    """)
    distributions = c.fetchall()
    
    # Convert to pandas DataFrame
    import pandas as pd
    from datetime import datetime
    
    df = pd.DataFrame(distributions, columns=[
        'رقم التوزيع', 'اسم المواطن', 'الرقم الوطني', 'الجوال', 'المادة',
        'الكمية', 'الوحدة', 'وزعت بواسطة', 'تاريخ التوزيع', 'ملاحظات'
    ])
    
    # Format dates and data
    if len(df) > 0:
        # Convert UTC timestamps to local timezone (12-hour format)
        local_tz = ZoneInfo(app.config['LOCAL_TIMEZONE'])
        df['تاريخ التوزيع'] = pd.to_datetime(df['تاريخ التوزيع'], utc=True).dt.tz_convert(local_tz).dt.strftime('%Y-%m-%d %I:%M %p').str.replace('AM', 'ص').str.replace('PM', 'م')
        df['ملاحظات'] = df['ملاحظات'].fillna('-')
        df['الكمية مع الوحدة'] = df['الكمية'].astype(str) + ' ' + df['الوحدة']
        
        # Reorder columns for better export
        df = df[['رقم التوزيع', 'اسم المواطن', 'الرقم الوطني', 'الجوال', 
                'المادة', 'الكمية مع الوحدة', 'وزعت بواسطة', 'تاريخ التوزيع', 'ملاحظات']]
    
    # Export to Excel
    from flask import make_response
    from io import BytesIO
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='سجل توزيع المواد', index=False)
    
    output.seek(0)
    
    response = make_response(output.read())
    # Use local time for filename timestamp
    local_tz = ZoneInfo(app.config['LOCAL_TIMEZONE'])
    local_now = datetime.now(timezone.utc).astimezone(local_tz)
    response.headers['Content-Disposition'] = f'attachment; filename=material_distributions_export_{local_now.strftime("%Y%m%d_%H%M%S")}.xlsx'
    response.headers['Content-type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    conn.close()
    
    return response

@app.route('/edit_distribution/<int:distribution_id>', methods=['GET', 'POST'])
@require_permission('distribute_materials')
def edit_distribution(distribution_id):
    conn = get_db_connection()
    c = conn.cursor()
    
    if request.method == 'POST':
        try:
            material_id = request.form.get('material_id')
            quantity = request.form.get('quantity')
            notes = request.form.get('notes', '')
            
            if not material_id or not quantity:
                flash('جميع الحقول المطلوبة يجب ملؤها', 'error')
                return redirect(request.url)
            
            # Update the distribution
            c.execute("""
                UPDATE material_distributions 
                SET material_id = ?, quantity = ?, notes = ?
                WHERE id = ?
            """, (material_id, quantity, notes, distribution_id))
            conn.commit()
            
            flash('تم تحديث التوزيع بنجاح', 'success')
            conn.close()
            return redirect(url_for('material_distributions'))
            
        except Exception as e:
            flash('خطأ في تحديث التوزيع', 'error')
            conn.close()
            return redirect(url_for('material_distributions'))
    
    # Get distribution details
    c.execute("""
        SELECT md.id, md.citizen_id, c.full_name, c.national_id, 
               md.material_id, m.name, md.quantity, md.notes
        FROM material_distributions md
        JOIN citizens c ON md.citizen_id = c.id
        JOIN materials m ON md.material_id = m.id
        WHERE md.id = ?
    """, (distribution_id,))
    distribution = c.fetchone()
    
    if not distribution:
        flash('التوزيع غير موجود', 'error')
        conn.close()
        return redirect(url_for('material_distributions'))
    
    # Get active materials
    c.execute("SELECT id, name, unit FROM materials WHERE is_active = 1 ORDER BY name")
    materials = c.fetchall()
    
    conn.close()
    
    return render_template('edit_distribution.html', distribution=distribution, materials=materials)

@app.route('/delete_distribution/<int:distribution_id>')
@require_permission('delete_materials')
def delete_distribution(distribution_id):
    conn = get_db_connection()
    c = conn.cursor()
    
    try:
        # Check if distribution exists
        c.execute("SELECT id FROM material_distributions WHERE id = ?", (distribution_id,))
        if not c.fetchone():
            flash('التوزيع غير موجود', 'error')
            conn.close()
            return redirect(url_for('material_distributions'))
        
        # Delete the distribution
        c.execute("DELETE FROM material_distributions WHERE id = ?", (distribution_id,))
        conn.commit()
        
        flash('تم حذف التوزيع بنجاح', 'success')
        
    except Exception as e:
        flash('خطأ في حذف التوزيع', 'error')
    
    conn.close()
    return redirect(url_for('material_distributions'))

@app.route('/export_materials')
@require_permission('view_materials')
def export_materials():
    conn = get_db_connection()
    c = conn.cursor()
    
    # Get all materials with user info
    c.execute("""
        SELECT m.id, m.name, m.description, m.unit, 
               CASE WHEN m.is_active = 1 THEN 'نشطة' ELSE 'معطلة' END as status,
               u.username, m.created_at
        FROM materials m
        JOIN users u ON m.created_by = u.id
        ORDER BY m.created_at DESC
    """)
    materials = c.fetchall()
    
    # Convert to pandas DataFrame
    import pandas as pd
    from datetime import datetime
    
    df = pd.DataFrame(materials, columns=[
        'رقم المادة', 'اسم المادة', 'الوصف', 'الوحدة', 
        'الحالة', 'أضيفت بواسطة', 'تاريخ الإضافة'
    ])
    
    # Format dates
    if len(df) > 0:
        # Convert UTC timestamps to local timezone (12-hour format)
        local_tz = ZoneInfo(app.config['LOCAL_TIMEZONE'])
        df['تاريخ الإضافة'] = pd.to_datetime(df['تاريخ الإضافة'], utc=True).dt.tz_convert(local_tz).dt.strftime('%Y-%m-%d %I:%M %p').str.replace('AM', 'ص').str.replace('PM', 'م')
        
        # Fill empty descriptions
        df['الوصف'] = df['الوصف'].fillna('-')
    
    # Export to Excel
    from flask import make_response
    from io import BytesIO
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='قائمة المواد', index=False)
    
    output.seek(0)
    
    response = make_response(output.read())
    # Use local time for filename timestamp
    local_tz = ZoneInfo(app.config['LOCAL_TIMEZONE'])
    local_now = datetime.now(timezone.utc).astimezone(local_tz)
    response.headers['Content-Disposition'] = f'attachment; filename=materials_export_{local_now.strftime("%Y%m%d_%H%M%S")}.xlsx'
    response.headers['Content-type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    conn.close()
    
    return response

@app.route('/export_citizen_materials/<int:citizen_id>')
@require_permission('view_distributions')
def export_citizen_materials(citizen_id):
    conn = get_db_connection()
    c = conn.cursor()
    
    # Get citizen information
    c.execute("SELECT full_name, national_id FROM citizens WHERE id = ?", (citizen_id,))
    citizen = c.fetchone()
    
    if not citizen:
        conn.close()
        flash('المواطن غير موجود', 'error')
        return redirect(url_for('view_citizens'))
    
    # Get material distributions for this citizen
    c.execute("""
        SELECT m.name, md.quantity, m.unit, md.distribution_date, u.username, md.notes
        FROM material_distributions md
        JOIN materials m ON md.material_id = m.id
        JOIN users u ON md.distributed_by = u.id
        WHERE md.citizen_id = ?
        ORDER BY md.distribution_date DESC
    """, (citizen_id,))
    distributions = c.fetchall()
    
    conn.close()
    
    if not distributions:
        flash('لا توجد مواد مستلمة لهذا المواطن', 'error')
        return redirect(url_for('view_citizens'))
    
    # Create DataFrame
    import pandas as pd
    df = pd.DataFrame(distributions, columns=[
        'اسم المادة', 'الكمية', 'الوحدة', 'تاريخ الاستلام', 'المسلم بواسطة', 'ملاحظات'
    ])
    
    # Format dates and data
    if len(df) > 0:
        # Convert UTC timestamps to local timezone (12-hour format)
        local_tz = ZoneInfo(app.config['LOCAL_TIMEZONE'])
        df['تاريخ الاستلام'] = pd.to_datetime(df['تاريخ الاستلام'], utc=True).dt.tz_convert(local_tz).dt.strftime('%Y-%m-%d %I:%M %p').str.replace('AM', 'ص').str.replace('PM', 'م')
        df['ملاحظات'] = df['ملاحظات'].fillna('-')
        df['الكمية مع الوحدة'] = df['الكمية'].astype(str) + ' ' + df['الوحدة']
        
        # Reorder columns for better export
        df = df[['اسم المادة', 'الكمية مع الوحدة', 'تاريخ الاستلام', 'المسلم بواسطة', 'ملاحظات']]
    
    # Export to Excel
    from flask import make_response
    from io import BytesIO
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='سجل المواد', index=False)
        
        # Get worksheet to add citizen info
        worksheet = writer.sheets['سجل المواد']
        
        # Insert rows at the top for citizen info
        worksheet.insert_rows(1, 3)
        worksheet['A1'] = f'الاسم: {citizen[0]}'
        worksheet['A2'] = f'الرقم الوطني: {citizen[1]}'
        # A3 will be empty for spacing
    
    output.seek(0)
    
    response = make_response(output.read())
    # Use local time for filename timestamp
    local_tz = ZoneInfo(app.config['LOCAL_TIMEZONE'])
    local_now = datetime.now(timezone.utc).astimezone(local_tz)
    # Use only ASCII characters for filename to avoid encoding issues
    import urllib.parse
    safe_name = f"citizen_{citizen_id}"  # Use citizen_id (integer) instead of name/national_id
    filename = f"{safe_name}_materials_{local_now.strftime('%Y%m%d_%H%M%S')}.xlsx"
    # URL encode filename to handle any special characters
    encoded_filename = urllib.parse.quote(filename)
    response.headers['Content-Disposition'] = f'attachment; filename="{encoded_filename}"'
    response.headers['Content-type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    return response

@app.route('/export_pdf')
def export_pdf():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    c.execute("SELECT * FROM citizens ORDER BY created_at DESC")
    citizens = c.fetchall()
    conn.close()
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4)
    
    # Create table data
    data = [['الرقم', 'الاسم', 'الرقم الوطني', 'الجوال', 'الحالة', 'أفراد الأسرة']]
    for citizen in citizens:
        data.append([
            str(citizen[0]),
            citizen[1][:20],
            citizen[2],
            citizen[3],
            citizen[4],
            str(citizen[5])
        ])
    
    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    doc.build([table])
    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name=f'citizens_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf',
        mimetype='application/pdf'
    )

# Database Reset Routes
@app.route('/database_reset')
@require_login
def database_reset():
    """صفحة إدارة تصفير قاعدة البيانات"""
    # فحص الصلاحيات - يجب أن يكون لديه إحدى صلاحيات التصفير
    reset_permissions = ['reset_citizens_data', 'reset_users_data', 'reset_materials_data', 'reset_all_data']
    has_any_reset_permission = any(has_permission(session['user_id'], perm) for perm in reset_permissions)
    
    if not has_any_reset_permission:
        flash('ليس لديك صلاحية للوصول لهذه الصفحة', 'error')
        return redirect(url_for('dashboard'))
    
    return render_template('database_reset.html')

@app.route('/reset_citizens_data', methods=['POST'])
@require_permission('reset_citizens_data')
@csrf_protect
def reset_citizens_data():
    """تصفير بيانات المواطنين"""
    
    try:
        conn = get_db_connection()
        c = conn.cursor()
        
        # حذف بيانات الحقول الديناميكية أولاً
        c.execute("DELETE FROM dynamic_field_values")
        # حذف توزيع المواد للمواطنين
        c.execute("DELETE FROM material_distributions")
        # حذف بيانات المواطنين
        c.execute("DELETE FROM citizens")
        
        conn.commit()
        conn.close()
        
        flash('تم تصفير بيانات المواطنين بنجاح', 'success')
    except Exception as e:
        flash(f'خطأ في تصفير بيانات المواطنين: {str(e)}', 'error')
    
    return redirect(url_for('database_reset'))

@app.route('/reset_users_data', methods=['POST'])
@require_permission('reset_users_data')
@csrf_protect
def reset_users_data():
    """تصفير بيانات المستخدمين (عدا الإدمن الأساسي)"""
    
    try:
        # التأكد من وجود إدمن أساسي قبل الحذف
        conn = get_db_connection()
        c = conn.cursor()
        c.execute("SELECT COUNT(*) FROM users WHERE is_admin = 1")
        admin_count = c.fetchone()[0]
        if admin_count < 1:
            flash('خطأ: لا يوجد مدير أساسي في النظام. العملية ملغاة لحماية النظام.', 'error')
            conn.close()
            return redirect(url_for('database_reset'))
        
        # حذف صلاحيات المستخدمين أولاً
        c.execute("DELETE FROM user_permissions WHERE user_id NOT IN (SELECT id FROM users WHERE username = 'admin')")
        # حذف المستخدمين (عدا الإدمن الأساسي)
        c.execute("DELETE FROM users WHERE username != 'admin'")
        
        conn.commit()
        conn.close()
        
        flash('تم تصفير بيانات المستخدمين بنجاح (تم الاحتفاظ بالإدمن الأساسي)', 'success')
    except Exception as e:
        flash(f'خطأ في تصفير بيانات المستخدمين: {str(e)}', 'error')
    
    return redirect(url_for('database_reset'))

@app.route('/reset_materials_data', methods=['POST'])
@require_permission('reset_materials_data')
@csrf_protect
def reset_materials_data():
    """تصفير بيانات المواد"""
    
    try:
        conn = get_db_connection()
        c = conn.cursor()
        
        # حذف توزيع المواد أولاً
        c.execute("DELETE FROM material_distributions")
        # حذف المواد
        c.execute("DELETE FROM materials")
        
        conn.commit()
        conn.close()
        
        flash('تم تصفير بيانات المواد بنجاح', 'success')
    except Exception as e:
        flash(f'خطأ في تصفير بيانات المواد: {str(e)}', 'error')
    
    return redirect(url_for('database_reset'))

@app.route('/reset_all_data', methods=['POST'])
@require_permission('reset_all_data')
@csrf_protect
def reset_all_data():
    """تصفير جميع البيانات عدا الإدمن الأساسي"""
    
    # إضافة تأكيد إضافي للعملية الخطيرة
    confirmation = request.form.get('confirmation', '').strip()
    if confirmation != 'تأكيد التصفير':
        flash('يجب كتابة "تأكيد التصفير" بدقة لتنفيذ هذه العملية', 'error')
        return redirect(url_for('database_reset'))
    
    try:
        # التأكد من وجود إدمن أساسي قبل الحذف  
        conn = get_db_connection()
        c = conn.cursor()
        c.execute("SELECT COUNT(*) FROM users WHERE is_admin = 1 AND username = 'admin'")
        admin_count = c.fetchone()[0]
        if admin_count < 1:
            flash('خطأ: لا يوجد مدير أساسي في النظام. العملية ملغاة لحماية النظام.', 'error')
            conn.close()
            return redirect(url_for('database_reset'))
        
        # حذف البيانات بالترتيب الصحيح لتجنب مشاكل المفاتيح الخارجية
        c.execute("DELETE FROM dynamic_field_values")
        c.execute("DELETE FROM material_distributions")
        c.execute("DELETE FROM citizens")
        c.execute("DELETE FROM materials")
        c.execute("DELETE FROM user_permissions WHERE user_id NOT IN (SELECT id FROM users WHERE username = 'admin')")
        c.execute("DELETE FROM users WHERE username != 'admin'")
        
        conn.commit()
        conn.close()
        
        flash('تم تصفير جميع البيانات بنجاح (تم الاحتفاظ بالإدمن الأساسي)', 'success')
    except Exception as e:
        flash(f'خطأ في تصفير البيانات: {str(e)}', 'error')
    
    return redirect(url_for('database_reset'))

# Excel Import Routes
@app.route('/import_citizens')
@require_permission('import_citizens_excel')
def import_citizens():
    """صفحة استيراد بيانات المواطنين من Excel"""
    # الحصول على قائمة المستخدمين لتحديد من سيُنسب إليه البيانات
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("SELECT id, username FROM users ORDER BY username")
    users = c.fetchall()
    conn.close()
    
    return render_template('import_citizens.html', users=users)

@app.route('/import_citizens_process', methods=['POST'])
@require_permission('import_citizens_excel')
@csrf_protect
def import_citizens_process():
    """معالجة استيراد بيانات المواطنين من Excel"""
    
    if 'excel_file' not in request.files:
        flash('لم يتم اختيار ملف Excel', 'error')
        return redirect(url_for('import_citizens'))
    
    file = request.files['excel_file']
    if file.filename == '':
        flash('لم يتم اختيار ملف', 'error')
        return redirect(url_for('import_citizens'))
    
    # التحقق من امتداد الملف ونوع المحتوى
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        flash('يجب أن يكون الملف من نوع Excel (.xlsx أو .xls)', 'error')
        return redirect(url_for('import_citizens'))
    
    # التحقق من نوع المحتوى
    allowed_mimes = [
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',  # .xlsx
        'application/vnd.ms-excel'  # .xls
    ]
    if file.content_type not in allowed_mimes:
        flash('نوع ملف غير مسموح. يجب أن يكون ملف Excel صحيح', 'error')
        return redirect(url_for('import_citizens'))
    
    assigned_user_id = request.form.get('assigned_user')
    if not assigned_user_id:
        flash('يجب اختيار المستخدم الذي ستُنسب إليه البيانات', 'error')
        return redirect(url_for('import_citizens'))
    
    try:
        # التحقق من وجود المستخدم
        conn = get_db_connection()
        c = conn.cursor()
        c.execute("SELECT username FROM users WHERE id = ?", (assigned_user_id,))
        user = c.fetchone()
        if not user:
            flash('المستخدم المحدد غير موجود', 'error')
            return redirect(url_for('import_citizens'))
        
        assigned_username = user[0]
        
        # قراءة ملف Excel مع حماية إضافية
        try:
            df = pd.read_excel(file, nrows=10000)  # حد أقصى 10000 سطر لحماية الذاكرة
        except Exception as e:
            flash(f'خطأ في قراءة ملف Excel: ملف تالف أو غير صحيح', 'error')
            conn.close()
            return redirect(url_for('import_citizens'))
        
        # التحقق من وجود الأعمدة المطلوبة
        required_columns = ['الاسم الثلاثي', 'الرقم الوطني', 'الجوال', 'الحالة', 'أفراد الأسرة', 'العنوان']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            flash(f'الأعمدة التالية مفقودة في ملف Excel: {", ".join(missing_columns)}', 'error')
            return redirect(url_for('import_citizens'))
        
        # تنظيف البيانات وإضافتها
        success_count = 0
        error_count = 0
        error_details = []
        
        for index, row in df.iterrows():
            try:
                # تنظيف البيانات
                full_name = str(row['الاسم الثلاثي']).strip()
                national_id = str(row['الرقم الوطني']).strip()
                phone = str(row['الجوال']).strip()
                status = str(row['الحالة']).strip()
                family_members = int(float(row['أفراد الأسرة']))
                address = str(row['العنوان']).strip()
                notes = str(row.get('الملاحظات', '')).strip() if pd.notna(row.get('الملاحظات')) else ''
                
                # تنسيق تلقائي للأرقام
                # إذا كان رقم الجوال 9 أرقام، أضف 0 في البداية ليصبح 10 أرقام
                if len(phone) == 9 and phone.isdigit():
                    phone = '0' + phone
                
                # إذا كان الرقم الوطني 10 أرقام، أضف 0 في البداية ليصبح 11 رقم
                if len(national_id) == 10 and national_id.isdigit():
                    national_id = '0' + national_id
                
                # التحقق من صحة البيانات
                if len(national_id) != 11 or not national_id.isdigit():
                    error_details.append(f'السطر {index + 2}: الرقم الوطني غير صحيح ({national_id})')
                    error_count += 1
                    continue
                
                if len(phone) != 10 or not phone.startswith('09') or not phone.isdigit():
                    error_details.append(f'السطر {index + 2}: رقم الجوال غير صحيح ({phone})')
                    error_count += 1
                    continue
                
                # التحقق من عدم وجود الرقم الوطني مسبقاً
                c.execute("SELECT id FROM citizens WHERE national_id = ?", (national_id,))
                if c.fetchone():
                    error_details.append(f'السطر {index + 2}: الرقم الوطني موجود مسبقاً ({national_id})')
                    error_count += 1
                    continue
                
                # إضافة البيانات
                c.execute("""
                    INSERT INTO citizens (full_name, national_id, phone, status, family_members, address, notes, added_by)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (full_name, national_id, phone, status, family_members, address, notes, assigned_username))
                
                success_count += 1
                
            except Exception as e:
                error_details.append(f'السطر {index + 2}: خطأ في المعالجة - {str(e)}')
                error_count += 1
        
        conn.commit()
        conn.close()
        
        # عرض النتائج
        if success_count > 0:
            flash(f'تم استيراد {success_count} سجل بنجاح', 'success')
        
        if error_count > 0:
            flash(f'فشل في استيراد {error_count} سجل', 'warning')
            # عرض أول 5 أخطاء فقط لتجنب ازدحام الرسائل
            for error in error_details[:5]:
                flash(error, 'error')
            if len(error_details) > 5:
                flash(f'و {len(error_details) - 5} أخطاء أخرى...', 'error')
        
    except Exception as e:
        flash(f'خطأ في قراءة ملف Excel: {str(e)}', 'error')
        conn.close()
    
    return redirect(url_for('import_citizens'))

@app.route('/download_excel_template')
@require_permission('import_citizens_excel')
def download_excel_template():
    """تحميل نموذج Excel لاستيراد بيانات المواطنين"""
    # إنشاء نموذج Excel
    data = {
        'الاسم الثلاثي': ['أحمد محمد علي', 'فاطمة أحمد محمد'],
        'الرقم الوطني': ['12345678901', '12345678902'],
        'الجوال': ['0912345678', '0987654321'],
        'الحالة': ['ارملة', 'حالة صعبة'],
        'أفراد الأسرة': [5, 3],
        'العنوان': ['غزة - الشجاعية', 'خان يونس - المواصي'],
        'الملاحظات': ['بحاجة لمساعدة عاجلة', 'حالة خاصة']
    }
    
    df = pd.DataFrame(data)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='نموذج البيانات', index=False)
        
        # تنسيق الملف
        worksheet = writer.sheets['نموذج البيانات']
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # تنسيق العناوين
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # تعديل عرض الأعمدة
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name='نموذج_استيراد_المواطنين.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

if __name__ == '__main__':
    init_db()
    app.run(host='0.0.0.0', port=5000, debug=True)
